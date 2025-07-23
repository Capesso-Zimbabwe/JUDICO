from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse, HttpResponseForbidden
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.admin.views.decorators import staff_member_required
import json
from django.utils import timezone
from datetime import timedelta, datetime
from django.db.models import Count, Q
from django.views.decorators.http import require_http_methods
from openpyxl import load_workbook
from io import BytesIO
import pandas as pd
import logging
from django.template.loader import render_to_string
from weasyprint import HTML, CSS
from django.conf import settings
import os
import uuid
from .models import Document, KYCBusiness, KYCProfile, KYCReport, KYCTestResult, KYCWorkflowState
from .forms import KYCBusinessForm, KYCProfileForm
from .services import KYCScreeningService
from .perform_kyc_screening import perform_kyc_screening
from django.core.files.base import ContentFile
import tempfile
from django.urls import reverse
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.db import transaction
from django.db import models

# Configure logger
logger = logging.getLogger(__name__)

# Create your views here.

@csrf_exempt
def register_kyc_business(request):
    """
    View to handle business KYC registration with multi-step form.
    """
    if request.method == "POST":
        try:
            # Check if this is a standard form submission or a multi-step submission
            if request.headers.get('Content-Type') == 'application/json':
                # Handle JSON submission for the original single-page form
                data = json.loads(request.body)
                form_data = data.get('form_data', {})
                owners_data = data.get('owners_data', [])

                # Create business profile
                business_form = KYCBusinessForm(form_data)
                if business_form.is_valid():
                    business = business_form.save()

                    # Add beneficial owners
                    for owner_data in owners_data:
                        try:
                            business.add_beneficial_owner(owner_data)
                        except Exception as e:
                            # If there's an error adding an owner, delete the business and return error
                            business.delete()
                            return JsonResponse({
                                'status': 'error',
                                'message': f'Error adding beneficial owner: {str(e)}'
                            })

                    return JsonResponse({
                        'status': 'success',
                        'message': 'Business KYC profile registered successfully'
                    })
                else:
                    return JsonResponse({
                        'status': 'error',
                        'message': 'Invalid business data',
                        'errors': business_form.errors
                    })
            else:
                # Handle multi-step form POST
                # Check if this is a draft or final submission
                is_draft = request.POST.get('is_draft') == 'true'
                current_step = int(request.POST.get('current_step', 1))
                
                # Process the form
                form = KYCBusinessForm(request.POST, request.FILES)
                
                # Get beneficial owners data
                beneficial_owners_json = request.POST.get('beneficial_owners_json', '[]')
                try:
                    beneficial_owners = json.loads(beneficial_owners_json)
                except json.JSONDecodeError:
                    beneficial_owners = []
                
                if form.is_valid():
                    # Save the business
                    business = form.save(commit=False)
                    business.is_draft = is_draft
                    business.save()
                    
                    # Add beneficial owners
                    update_beneficial_owners(business, beneficial_owners)
                    
                    # If this is a final submission, update the workflow state
                    if not is_draft:
                        try:
                            # Create workflow state or transition to SUBMITTED
                            workflow_state = create_workflow_state(
                                business.id, 'business', 'SUBMITTED'
                            )
                            # Import any documents to Document model
                            import_id_document(business=business)
                        except Exception as e:
                            messages.warning(request, f"Business profile saved but workflow state update failed: {str(e)}")
                    
                    # Set success message based on draft status
                    if is_draft:
                        messages.success(request, "Business KYC profile saved as draft. You can complete it later.")
                    else:
                        messages.success(request, "Business KYC profile submitted successfully!")
                    
                    # Redirect based on draft status
                    if is_draft:
                        return redirect('kyc_app:kyc_business_drafts')
                    else:
                        return redirect('kyc_app:kyc_workflow_dashboard')
                else:
                    # Show form errors
                    messages.error(request, "There were errors in your submission. Please correct them and try again.")

        except Exception as e:
            messages.error(request, f"An error occurred: {str(e)}")
            # For non-JSON requests, return to the form page
            # For REST API clients, still preparing a JsonResponse
            if request.headers.get('Content-Type') == 'application/json':
                return JsonResponse({
                    'status': 'error',
                    'message': str(e)
                })
    
    # For GET requests or if form validation fails
    # Check if we are resuming a draft
    draft_id = request.GET.get('resume_draft')
    if draft_id:
        try:
            # Try to find the draft business
            draft_business = KYCBusiness.objects.get(id=draft_id, is_draft=True)
            # Initialize form with the draft data
            form = KYCBusinessForm(instance=draft_business)
            
            # Get beneficial owners
            beneficial_owners = KYCBeneficialOwner.objects.filter(business=draft_business)
            beneficial_owners_json = json.dumps([
                {
                    'name': owner.name,
                    'nationality': owner.nationality,
                    'id_document_type': owner.id_document_type,
                    'id_document_number': owner.id_document_number,
                    'ownership_percentage': float(owner.ownership_percentage),
                    'pep_status': owner.pep_status
                } for owner in beneficial_owners
            ])
            
            messages.info(request, "You're editing a draft Business KYC profile. Complete and submit when ready.")
        except KYCBusiness.DoesNotExist:
            messages.error(request, "Draft business profile not found.")
            form = KYCBusinessForm()
            beneficial_owners_json = '[]'
    else:
        # Regular new form
        form = KYCBusinessForm()
        beneficial_owners_json = '[]'
    
    # Get user's draft businesses for the dropdown
    if request.user.is_authenticated:
        draft_businesses = KYCBusiness.objects.filter(is_draft=True)
    else:
        draft_businesses = []
    
    # Choose template for multi-step form
    template_name = 'register_kyc_business_multi.html'
    
    return render(request, template_name, {
        "form": form,
        "draft_businesses": draft_businesses,
        "beneficial_owners_json": beneficial_owners_json
    })

@login_required
def register_kyc_profile(request):
    """
    View to register a new KYC Profile with support for multi-step form and drafts.
    """
    if request.method == "POST":
        # Check if this is a draft or final submission
        is_draft = request.POST.get('is_draft') == 'true'
        current_step = int(request.POST.get('current_step', 1))
        
        # Process the form
        form = KYCProfileForm(request.POST, request.FILES)
        
        if form.is_valid():
            # Save the profile
            profile = form.save(commit=False)
            profile.is_draft = is_draft
            profile.save()
            
            # If this is a final submission, update the workflow state
            if not is_draft:
                try:
                    # Transition to SUBMITTED state
                    workflow_state = profile.workflow_state
                    workflow_state.transition_to(
                        'SUBMITTED',
                        user=request.user.username if request.user.is_authenticated else 'Anonymous',
                        notes='Submitted via KYC registration form'
                    )
                except Exception as e:
                    messages.warning(request, f"Profile saved but workflow state update failed: {str(e)}")
            
            # Set success message based on draft status
            if is_draft:
                messages.success(request, "KYC profile saved as draft. You can complete it later.")
            else:
                messages.success(request, "KYC profile submitted successfully!")
            
            # Reinitialize form to clear the fields after a successful submission
            form = KYCProfileForm()
        else:
            # Show form errors
            messages.error(request, "There were errors in your submission. Please correct them and try again.")
    else:
        # For GET requests, check if we are resuming a draft
        draft_id = request.GET.get('resume_draft')
        if draft_id:
            try:
                # Try to find the draft profile
                draft_profile = KYCProfile.objects.get(id=draft_id, is_draft=True)
                # Initialize form with the draft data
                form = KYCProfileForm(instance=draft_profile)
                messages.info(request, "You're editing a draft KYC profile. Complete and submit when ready.")
            except KYCProfile.DoesNotExist:
                messages.error(request, "Draft profile not found.")
                form = KYCProfileForm()
        else:
            # Regular new form
            form = KYCProfileForm()
    
    # Get user's draft profiles for the dropdown
    if request.user.is_authenticated:
        # In a real app, you'd filter by the current user
        draft_profiles = KYCProfile.objects.filter(is_draft=True)
    else:
        draft_profiles = []
    
    # Choose template based on preference - standard or multi-step
    # You can make this configurable or use A/B testing to see which works better
    template_name = 'register_kyc_profile_multi.html'  # For the multi-step form
    # template_name = 'register_kyc_profile.html'  # For the standard form
    
    return render(request, template_name, {
        "form": form,
        "draft_profiles": draft_profiles
    })

@login_required
def kyc_profile_drafts(request):
    """
    View to list all draft KYC profiles.
    """
    # In a real app, you'd filter by the current user
    draft_profiles = KYCProfile.objects.filter(is_draft=True)
    
    return render(request, "kyc_profile_drafts.html", {
        "draft_profiles": draft_profiles
    })

@login_required
def delete_kyc_draft(request, profile_id):
    """
    View to delete a draft KYC profile.
    """
    profile = get_object_or_404(KYCProfile, id=profile_id, is_draft=True)
    
    if request.method == "POST":
        profile.delete()
        messages.success(request, "Draft profile deleted successfully.")
        return redirect('kyc_app:kyc_profile_drafts')
    
    return render(request, "confirm_delete_draft.html", {
        "profile": profile
    })

@login_required
def kyc_workflow_dashboard(request):
    """
    View for the KYC workflow dashboard showing all profiles and businesses in their various states.
    """
    # Get individual KYC profiles by workflow state
    draft_profiles = KYCProfile.objects.filter(is_draft=True)
    
    # Get business KYC profiles by workflow state
    draft_businesses = KYCBusiness.objects.filter(is_draft=True)
    
    # Get workflow states for both individuals and businesses
    submitted_profiles = KYCWorkflowState.objects.filter(current_state='SUBMITTED')
    
    # Make sure to include APPROVAL_PENDING in the in_review_states
    in_review_states = ['DOC_REVIEW', 'SCREENING', 'INVESTIGATION', 'APPROVAL_PENDING']
    in_review_profiles = KYCWorkflowState.objects.filter(current_state__in=in_review_states)
    
    # Ensure we're getting all approved and rejected profiles
    approved_profiles = KYCWorkflowState.objects.filter(current_state='APPROVED')
    rejected_profiles = KYCWorkflowState.objects.filter(current_state='REJECTED')
    
    # Debug: Print counts to console
    print(f"DEBUG: Draft profiles: {draft_profiles.count()}")
    print(f"DEBUG: Submitted profiles: {submitted_profiles.count()}")
    print(f"DEBUG: In review profiles: {in_review_profiles.count()}")
    print(f"DEBUG: Approved profiles: {approved_profiles.count()}")
    print(f"DEBUG: Rejected profiles: {rejected_profiles.count()}")
    
    # Calculate dashboard metrics
    today = timezone.now().date()
    week_ago = today - timedelta(days=7)
    
    # Total profiles and businesses
    total_individual_profiles = KYCProfile.objects.count()
    total_business_profiles = KYCBusiness.objects.count()
    total_profiles = total_individual_profiles + total_business_profiles
    
    # Active profiles (those not in draft state)
    active_profiles = total_profiles - draft_profiles.count() - draft_businesses.count()
    if active_profiles < 0:
        active_profiles = 0
    
    # New profiles this week (both individuals and businesses)
    new_individual_profiles = KYCProfile.objects.filter(created_at__gte=week_ago).count()
    new_business_profiles = KYCBusiness.objects.filter(created_at__gte=week_ago).count()
    new_profiles_count = new_individual_profiles + new_business_profiles
    
    # Profiles with expiring documents in the next 30 days (only for individuals)
    thirty_days_from_now = today + timedelta(days=30)
    expiring_docs = KYCProfile.objects.filter(
        id_expiry_date__lte=thirty_days_from_now,
        id_expiry_date__gt=today
    ).count()
    
    # Already expired documents
    expired_docs = KYCProfile.objects.filter(
        id_expiry_date__lt=today
    ).count()
    
    # Pending review counts
    pending_reviews = submitted_profiles.count() + sum(
        KYCWorkflowState.objects.filter(current_state=state).count() 
        for state in ['DOC_REVIEW', 'APPROVAL_PENDING']
    )
    
    # High risk pending review
    high_risk_pending = KYCTestResult.objects.filter(
        risk_level='High',
        kyc_profile__workflow_state__current_state__in=['SUBMITTED'] + in_review_states
    ).count()
    
    # Risk metrics
    test_results = KYCTestResult.objects.all()
    high_risk_profiles = test_results.filter(risk_level='High').count()
    medium_risk_count = test_results.filter(risk_level='Medium').count()
    low_risk_count = test_results.filter(risk_level='Low').count()
    
    total_risk_count = high_risk_profiles + medium_risk_count + low_risk_count
    if total_risk_count > 0:
        high_risk_percent = round((high_risk_profiles / total_risk_count) * 100)
        medium_risk_percent = round((medium_risk_count / total_risk_count) * 100)
        low_risk_percent = round((low_risk_count / total_risk_count) * 100)
    else:
        high_risk_percent = medium_risk_percent = low_risk_percent = 0
    
    # Compile recent activity from workflow state history
    recent_activities = []
    for state in KYCWorkflowState.objects.all().order_by('-updated_at')[:20]:
        # Extract activities from the history JSON
        if state.history and len(state.history) > 0:
            for activity in state.history[-3:]:  # Get the most recent 3 activities
                if 'timestamp' in activity and 'from_state' in activity and 'to_state' in activity:
                    try:
                        recent_activities.append({
                            'timestamp': timezone.datetime.fromisoformat(activity['timestamp']),
                            'profile_name': state.get_subject_name(),
                            'profile_id': state.get_subject_id(),
                            'description': f"Changed from {activity['from_state']} to {activity['to_state']}",
                            'status': 'Success',
                            'user': activity.get('by_user', 'System'),
                            'is_business': bool(state.business_kyc)
                        })
                    except (ValueError, TypeError):
                        # Skip if timestamp isn't valid
                        pass
    
    # Sort activities by timestamp
    recent_activities.sort(key=lambda x: x['timestamp'], reverse=True)
    recent_activities = recent_activities[:10]  # Limit to 10 most recent
    
    # Debug - help diagnose the issue
    print(f"DEBUG: Recent activities count: {len(recent_activities)}")
    
    # Adding some debug info to the context
    context = {
        # Profiles by workflow state
        'draft_profiles': draft_profiles,
        'draft_businesses': draft_businesses,
        'submitted_profiles': submitted_profiles,
        'review_profiles': in_review_profiles,  # Make sure this is consistent with template
        'in_review_profiles': in_review_profiles,  # Add this to ensure compatibility  
        'approved_profiles': approved_profiles,
        'rejected_profiles': rejected_profiles,
        
        # Dashboard metrics
        'total_profiles': total_profiles,
        'active_profiles': active_profiles,
        'total_individual_profiles': total_individual_profiles,
        'total_business_profiles': total_business_profiles,
        'new_profiles_count': new_profiles_count,
        'expiring_docs': expiring_docs,
        'expired_docs': expired_docs,
        'pending_reviews': pending_reviews,
        'high_risk_pending': high_risk_pending,
        'high_risk_profiles': high_risk_profiles,
        'high_risk_percent': high_risk_percent,
        'medium_risk_percent': medium_risk_percent,
        'low_risk_percent': low_risk_percent,
        
        # Recent activities
        'recent_activities': recent_activities,
        
        # Today for date comparisons in template
        'today': today,
        
        # Debug info
        'debug_submitted_count': submitted_profiles.count(),
        'debug_review_count': in_review_profiles.count(),
        'debug_approved_count': approved_profiles.count(),
        'debug_rejected_count': rejected_profiles.count(),
    }
    
    return render(request, "kyc_workflow_dashboard.html", context)

@login_required
def register_kyc_business_multi(request):
    """
    View for multi-step business KYC registration form.
    Handles both GET and POST requests for creating/updating business profile.
    """
    # Get the current user ID
    user_id = request.user.id if request.user.is_authenticated else None
    form = KYCBusinessForm()
    
    # For POST requests, process form submission
    if request.method == 'POST':
        # Get the draft ID if provided
        draft_id = request.POST.get('draft_id', None)
        
        # Handle document uploads
        registration_document = request.FILES.get('registration_document', None)
        tax_document = request.FILES.get('tax_document', None)
        
        # Check if this is a draft submission
        is_draft = request.POST.get('is_draft', 'false').lower() == 'true'
        
        # Get beneficial owners from JSON string
        beneficial_owners_json = request.POST.get('beneficial_owners_json', '[]')
        beneficial_owners = json.loads(beneficial_owners_json)
        
        form = KYCBusinessForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                # DEBUG: Print length of all fields with max_length=50 constraint
                print("DEBUG - Field lengths for fields with 50-char limit:")
                fifty_char_fields = [
                    'business_id', 'business_type', 'industry_sector', 
                    'ownership_structure', 'annual_revenue', 'source_of_funds',
                    'transaction_volume', 'account_number', 'account_type', 'swift_code'
                ]
                for field_name in fifty_char_fields:
                    field_value = request.POST.get(field_name, '')
                    field_length = len(field_value) if field_value else 0
                    print(f"  {field_name}: {field_length} chars - '{field_value}'")
                    if field_length > 50:
                        print(f"  WARNING: '{field_name}' exceeds 50 char limit: {field_length}")
                
                # Start a transaction
                with transaction.atomic():
                    # DEBUGGING: Log all field values and their lengths
                    print("\n\n==== DEBUGGING POST DATA ====")
                    print("Checking all field values for length issues:")
                    for field_name, field_value in request.POST.items():
                        field_length = len(str(field_value))
                        print(f"{field_name}: {field_length} chars")
                        if field_length > 50:
                            print(f"WARNING: Field '{field_name}' exceeds 50 chars (length={field_length})")
                        if field_length > 100:
                            print(f"CRITICAL: Field '{field_name}' exceeds 100 chars (length={field_length})")
                    print("==== END DEBUGGING ====\n\n")
                    
                    # Save the business profile
                    business_kyc = form.save(commit=False)
                    
                    # Set draft status
                    business_kyc.is_draft = is_draft
                    
                    # Save to generate ID if new
                    business_kyc.save()
                    
                    # If business is saved, create or update workflow state
                    workflow, created = KYCWorkflowState.objects.get_or_create(
                        business_kyc=business_kyc,
                        defaults={
                            'current_state': 'DRAFT' if is_draft else 'SUBMITTED',
                            'last_modified_by': f"User ID: {user_id}" if user_id else "Anonymous"
                        }
                    )
                    
                    # If existing workflow, update state if not draft
                    if not created and not is_draft:
                        workflow.current_state = 'SUBMITTED'
                        workflow.last_modified_by = f"User ID: {user_id}" if user_id else "Anonymous"
                        workflow.save()
                    
                    # Process beneficial owners
                    if beneficial_owners:
                        # For now, store as a text field
                        # Ideally, you would create a BeneficialOwner model linked to business_kyc
                        # And store each beneficial owner separately
                        business_kyc.ownership_structure = json.dumps(beneficial_owners)
                        business_kyc.save(update_fields=['ownership_structure'])
                    
                    # Import documents to Document model if needed
                    if business_kyc.registration_document or business_kyc.tax_document:
                        try:
                            import_id_document(business=business_kyc)
                        except Exception as e:
                            print(f"Error importing documents: {str(e)}")
                    
                    if is_draft:
                        messages.success(request, "Business KYC profile saved as draft successfully.")
                    else:
                        messages.success(request, "Business KYC profile submitted successfully. Awaiting review.")
                    
                    # Redirect to dashboard or appropriate page
                    return redirect('kyc_app:business_dashboard')  # Adjust as needed
                
            except Exception as e:
                error_message = str(e)
                print(f"Error saving business KYC: {error_message}")
                
                # More detailed error handling for common database errors
                if "character varying(50)" in error_message:
                    # This is the "value too long" error - try to identify the field
                    # Check each field with max_length=50 constraint
                    for field_name in ['business_id', 'business_type', 'industry_sector', 
                                       'ownership_structure', 'annual_revenue', 'source_of_funds',
                                       'transaction_volume', 'account_number', 'account_type', 'swift_code']:
                        field_value = request.POST.get(field_name, '')
                        if len(field_value) > 50:
                            field_label = form.fields[field_name].label or field_name.replace('_', ' ').title()
                            messages.error(request, f"Error in {field_label}: Value is too long (maximum 50 characters)")
                            # Add field-specific error to the form
                            form.add_error(field_name, "This field value is too long (maximum 50 characters)")
                            break
                    else:
                        # If no specific field found, show generic message
                        messages.error(request, f"Error saving business profile: A field value is too long (maximum 50 characters)")
                else:
                    # For other types of errors
                    messages.error(request, f"Error saving business profile: {error_message}")
        else:
            # If form is not valid, add form errors to messages
            for field, errors in form.errors.items():
                field_label = form.fields[field].label or field.replace('_', ' ').title()
                for error in errors:
                    messages.error(request, f"{field_label}: {error}")
    
    # For GET requests, prepare new form
    
    # Create empty list for owners if needed
    beneficial_owners_json = '[]'
    
    # Prepare context with form
    context = {
        'form': form,
        'beneficial_owners_json': beneficial_owners_json
    }
    
    return render(request, 'register_kyc_business_multi.html', context)

@login_required
def review_kyc_profile(request, profile_id):
    """
    View to review an individual KYC profile, perform KYC screening,
    and allow for approval or rejection based on the results.
    """
    profile = get_object_or_404(KYCProfile, id=profile_id)
    workflow_state = profile.workflow_state
    
    # Get all documents related to this profile
    documents = Document.objects.filter(profile=profile)
    
    # Count verified, pending, and rejected documents
    verified_documents_count = documents.filter(status='VERIFIED').count()
    pending_documents_count = documents.filter(status='PENDING').count()
    rejected_documents_count = documents.filter(status='REJECTED').count()
    
    # Check if all documents are verified
    all_documents_verified = (documents.count() > 0 and 
                             verified_documents_count == documents.count())
    
    # Flag if there are pending documents
    has_pending_documents = pending_documents_count > 0
    
    # Define required document types based on policy
    required_document_types = ['ID_DOCUMENT', 'PROOF_OF_ADDRESS']
    
    # Check which required documents are missing
    existing_document_types = list(documents.values_list('document_type', flat=True))
    missing_required_documents = [
        doc_type for doc_type in required_document_types 
        if doc_type not in existing_document_types
    ]
    
    # Convert document type codes to display names for better readability
    document_type_map = dict(Document.DOCUMENT_TYPE_CHOICES)
    missing_required_documents = [document_type_map.get(doc_type, doc_type) 
                                 for doc_type in missing_required_documents]
    
    # Check if this is a POST request (approval/rejection action)
    if request.method == "POST":
        action = request.POST.get('action')
        notes = request.POST.get('notes', '')
        
        if action == 'approve':
            # Only allow approval if all documents are verified
            if not all_documents_verified or missing_required_documents:
                messages.error(
                    request, 
                    "Cannot approve profile. All required documents must be uploaded and verified."
                )
                return redirect('kyc_app:review_kyc_profile', profile_id=profile.id)
            
            # Debug print to help with troubleshooting
            print(f"DEBUG: Approving profile {profile.id}, current state: {workflow_state.current_state}")
            
            # Directly update the workflow state for approval
            workflow_state.current_state = 'APPROVED'
            workflow_state.approved_by = request.user.username
            workflow_state.approval_date = timezone.now()
            workflow_state.reviewer_notes = notes or 'Manual approval after screening'
            workflow_state.next_review_date = (timezone.now() + timedelta(days=365)).date()
            
            # Update history
            history_entry = {
                'from_state': 'APPROVAL_PENDING',
                'to_state': 'APPROVED',
                'timestamp': timezone.now().isoformat(),
                'by_user': request.user.username,
                'notes': notes or 'Manual approval after screening'
            }
            
            # Add to history
            if not workflow_state.history:
                workflow_state.history = []
            workflow_state.history.append(history_entry)
            
            # Save changes
            workflow_state.save()
            
            print(f"DEBUG: After approval, state is now: {workflow_state.current_state}")
            
            messages.success(request, f"KYC profile for {profile.full_name} has been approved.")
            return redirect('kyc_app:kyc_workflow_dashboard')
            
        elif action == 'reject':
            reason = request.POST.get('rejection_reason', '')
            
            # Debug print to help with troubleshooting
            print(f"DEBUG: Rejecting profile {profile.id}, current state: {workflow_state.current_state}")
            
            # Directly update the workflow state for rejection
            workflow_state.current_state = 'REJECTED'
            workflow_state.rejection_reason = reason
            workflow_state.reviewer_notes = notes or 'Manual rejection after screening'
            
            # Update history
            history_entry = {
                'from_state': workflow_state.current_state,
                'to_state': 'REJECTED',
                'timestamp': timezone.now().isoformat(),
                'by_user': request.user.username,
                'notes': notes or 'Manual rejection after screening'
            }
            
            # Add to history
            if not workflow_state.history:
                workflow_state.history = []
            workflow_state.history.append(history_entry)
            
            # If rejection reason is missing documents, log which ones
            if reason == 'Missing Required Documents' and missing_required_documents:
                additional_notes = f"Missing documents: {', '.join(missing_required_documents)}"
                workflow_state.reviewer_notes = (workflow_state.reviewer_notes or '') + '\n' + additional_notes
            
            # Save changes
            workflow_state.save()
            
            print(f"DEBUG: After rejection, state is now: {workflow_state.current_state}")
            
            messages.success(request, f"KYC profile for {profile.full_name} has been rejected.")
            return redirect('kyc_app:kyc_workflow_dashboard')
    
    # Handle GET request - perform KYC screening
    
    # Transition to screening state
    workflow_state.transition_to(
        'SCREENING',
        user=request.user.username,
        notes='Manual screening initiated during review'
    )
    
    # Perform KYC screening
    test_result = perform_kyc_screening(profile.id_document_number)
    
    # Fetch latest test results after screening
    kyc_tests = KYCTestResult.objects.filter(kyc_profile=profile).order_by('-created_at')
    
    # After screening, transition to approval pending
    workflow_state.transition_to(
        'APPROVAL_PENDING',
        user=request.user.username,
        notes='Screening completed, awaiting approval decision'
    )
    
    # Check if we have a proper test result
    error_message = None
    if isinstance(test_result, str):
        error_message = test_result
        test_result = None
    
    return render(request, 'review_kyc_profile.html', {
        'profile': profile,
        'workflow_state': workflow_state,
        'test_result': test_result if not isinstance(test_result, str) else None,
        'kyc_tests': kyc_tests,
        'error_message': error_message,
        # Document verification status
        'documents': documents,
        'verified_documents_count': verified_documents_count,
        'pending_documents_count': pending_documents_count,
        'rejected_documents_count': rejected_documents_count,
        'all_documents_verified': all_documents_verified,
        'has_pending_documents': has_pending_documents,
        'missing_required_documents': missing_required_documents
    })

@login_required
def review_kyc_business(request, business_id):
    """
    View to review a business KYC profile, perform KYC screening,
    and allow for approval or rejection based on the results.
    """
    business = get_object_or_404(KYCBusiness, id=business_id)
    workflow_state = business.workflow_state
    
    # Get registration and tax documents status
    registration_doc = None
    tax_doc = None
    registration_doc_id = None
    tax_doc_id = None
    
    # For businesses, we use direct file fields rather than Document model
    # Business documents are stored directly on the KYCBusiness model
    documents = []  # Empty list instead of queryset since we don't use Document model for businesses
    
    # Check for registration document status
    has_verified_registration_doc = False
    has_rejected_registration_doc = False
    has_pending_registration_doc = False
    
    # Check for direct registration document file
    if business.registration_document:
        # For businesses, we just need to check if the file exists
        has_pending_registration_doc = True  # Assume pending by default
        # Registration doc exists but needs verification (we don't track verification status for direct uploads)
    
    # Check for tax document status
    has_verified_tax_doc = False
    has_rejected_tax_doc = False
    has_pending_tax_doc = False
    
    # Check for direct tax document file
    if business.tax_document:
        # For businesses, we just need to check if the file exists
        has_pending_tax_doc = True  # Assume pending by default
        # Tax doc exists but needs verification (we don't track verification status for direct uploads)
    
    # For businesses we don't have "additional" documents in the Document model
    additional_documents = []
    
    # Count verified, pending, and rejected documents
    verified_documents_count = 0  # No verified documents as we don't track verification
    
    pending_documents_count = (
        (1 if has_pending_registration_doc else 0) + 
        (1 if has_pending_tax_doc else 0)
    )
    
    rejected_documents_count = 0  # No rejected documents as we don't track rejection
    
    # Check if all required documents are present
    total_documents = 0
    if business.registration_document:
        total_documents += 1
    if business.tax_document:
        total_documents += 1
    
    # For business we'll consider the document verified if it's uploaded
    all_documents_verified = (
        total_documents >= 2  # Both registration and tax documents are required
    )
    
    # Flag if there are pending documents (for businesses, all uploaded docs are considered pending)
    has_pending_documents = pending_documents_count > 0
    
    # Define required document types for businesses
    required_document_types = ['BUSINESS_REGISTRATION', 'TAX_DOCUMENT']
    
    # Check which required documents are missing
    existing_document_types = []
    
    # Add registration document if it exists
    if business.registration_document:
        existing_document_types.append('BUSINESS_REGISTRATION')
    
    # Add tax document if it exists
    if business.tax_document:
        existing_document_types.append('TAX_DOCUMENT')
    
    missing_required_documents = [
        doc_type for doc_type in required_document_types
        if doc_type not in existing_document_types
    ]
    
    # Convert document type codes to display names for better readability
    document_type_map = {
        'BUSINESS_REGISTRATION': 'Business Registration',
        'TAX_DOCUMENT': 'Tax Document'
    }
    missing_required_documents = [document_type_map.get(doc_type, doc_type)
                                 for doc_type in missing_required_documents]
    
    # Check if this is a POST request (approval/rejection action)
    if request.method == "POST":
        action = request.POST.get('action')
        notes = request.POST.get('notes', '')
        
        if action == 'approve':
            # Only allow approval if all documents are verified
            if not all_documents_verified or missing_required_documents:
                messages.error(
                    request, 
                    "Cannot approve business. All required documents must be uploaded and verified."
                )
                return redirect('kyc_app:review_kyc_business', business_id=business.id)
                
            # Debug print to help with troubleshooting
            print(f"DEBUG: Approving business {business.id}, current state: {workflow_state.current_state}")
            
            # Directly update the workflow state for approval
            workflow_state.current_state = 'APPROVED'
            workflow_state.approved_by = request.user.username
            workflow_state.approval_date = timezone.now()
            workflow_state.reviewer_notes = notes or 'Manual approval after screening'
            workflow_state.next_review_date = (timezone.now() + timedelta(days=365)).date()
            
            # Update history
            history_entry = {
                'from_state': 'APPROVAL_PENDING',
                'to_state': 'APPROVED',
                'timestamp': timezone.now().isoformat(),
                'by_user': request.user.username,
                'notes': notes or 'Manual approval after screening'
            }
            
            # Add to history
            if not workflow_state.history:
                workflow_state.history = []
            workflow_state.history.append(history_entry)
            
            # Save changes
            workflow_state.save()
            
            print(f"DEBUG: After approval, state is now: {workflow_state.current_state}")
            
            messages.success(request, f"Business KYC profile for {business.business_name} has been approved.")
            return redirect('kyc_app:kyc_workflow_dashboard')
            
        elif action == 'reject':
            reason = request.POST.get('rejection_reason', '')
            
            # Debug print to help with troubleshooting
            print(f"DEBUG: Rejecting business {business.id}, current state: {workflow_state.current_state}")
            
            # Directly update the workflow state for rejection
            workflow_state.current_state = 'REJECTED'
            workflow_state.rejection_reason = reason
            workflow_state.reviewer_notes = notes or 'Manual rejection after screening'
            
            # Update history
            history_entry = {
                'from_state': workflow_state.current_state,
                'to_state': 'REJECTED',
                'timestamp': timezone.now().isoformat(),
                'by_user': request.user.username,
                'notes': notes or 'Manual rejection after screening'
            }
            
            # Add to history
            if not workflow_state.history:
                workflow_state.history = []
            workflow_state.history.append(history_entry)
            
            # If rejection reason is missing documents, log which ones
            if reason == 'Missing Required Documents' and missing_required_documents:
                additional_notes = f"Missing documents: {', '.join(missing_required_documents)}"
                workflow_state.reviewer_notes = (workflow_state.reviewer_notes or '') + '\n' + additional_notes
            
            # Save changes
            workflow_state.save()
            
            print(f"DEBUG: After rejection, state is now: {workflow_state.current_state}")
            
            messages.success(request, f"Business KYC profile for {business.business_name} has been rejected.")
            return redirect('kyc_app:kyc_workflow_dashboard')
    
    # Handle GET request - perform assessment
    
    # Transition to screening state
    workflow_state.transition_to(
        'SCREENING',
        user=request.user.username,
        notes='Manual risk assessment initiated during review'
    )
    
    # For businesses, we'll check for high-risk jurisdictions and other factors
    
    # After assessment, transition to approval pending
    workflow_state.transition_to(
        'APPROVAL_PENDING',
        user=request.user.username,
        notes='Risk assessment completed, awaiting approval decision'
    )
    
    return render(request, 'review_kyc_business.html', {
        'business': business,
        'workflow_state': workflow_state,
        'high_risk_jurisdiction': business.high_risk_jurisdiction,
        'beneficial_owners': business.get_beneficial_owners(),
        # Document verification status - empty lists since we don't use Document model for businesses
        'documents': [],
        'additional_documents': [],
        # Document status indicators
        'has_verified_registration_doc': False,  # Not tracked for direct uploads
        'has_rejected_registration_doc': False,  # Not tracked for direct uploads
        'has_pending_registration_doc': business.registration_document is not None,
        'has_verified_tax_doc': False,  # Not tracked for direct uploads
        'has_rejected_tax_doc': False,  # Not tracked for direct uploads
        'has_pending_tax_doc': business.tax_document is not None,
        # Document IDs (none for direct uploads)
        'registration_doc_id': None,
        'tax_doc_id': None,
        # Document counts
        'verified_documents_count': 0,
        'pending_documents_count': (1 if business.registration_document else 0) + (1 if business.tax_document else 0),
        'rejected_documents_count': 0,
        # Status flags
        'all_documents_verified': business.registration_document is not None and business.tax_document is not None,
        'has_pending_documents': business.registration_document is not None or business.tax_document is not None,
        'missing_required_documents': [
            doc for doc in [
                'Business Registration' if not business.registration_document else None,
                'Tax Document' if not business.tax_document else None
            ] if doc is not None
        ]
    })

@login_required
@require_http_methods(["GET", "POST"])
def bulk_import_profiles(request):
    """
    View to handle bulk import of KYC profiles and businesses from Excel files.
    This allows for mass upload of profiles which will then be processed through the KYC workflow.
    """
    # Store import results in session for showing after redirect
    import_results = {
        'total': 0,
        'success': 0,
        'failed': 0,
        'errors': [],
        'completed': False
    }
    
    if request.method == "POST" and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        import_type = request.POST.get('import_type', 'individual')  # Default to individual
        
        # Track import statistics
        stats = {
            'total': 0,
            'success': 0,
            'failed': 0,
            'errors': []
        }
        
        try:
            # Process using pandas for better error handling
            df = pd.read_excel(excel_file)
            
            # Log the columns found in the Excel file
            columns = list(df.columns)
            print(f"DEBUG: Excel file contains columns: {columns}")
            
            if import_type == 'individual':
                # Process individual profiles
                for idx, row in df.iterrows():
                    stats['total'] += 1
                    try:
                        print(f"DEBUG: Processing individual profile row {idx+1}: {row['full_name'] if 'full_name' in row else 'Unknown'}")
                        
                        # Create profile from row data
                        data = {
                            'customer_id': row.get('customer_id', f"IMP-{stats['total']}"),
                            'full_name': row['full_name'],
                            'date_of_birth': row.get('date_of_birth'),
                            'nationality': row.get('nationality', 'Unknown'),
                            'id_document_type': row.get('id_document_type', 'Passport'),
                            'id_document_number': row.get('id_document_number', ''),
                            'id_issued_country': row.get('id_issued_country', ''),
                            'email': row.get('email', ''),
                            'phone_number': row.get('phone_number', ''),
                            'address': row.get('address', ''),
                            'city': row.get('city', ''),
                            'country': row.get('country', ''),
                            'account_number': row.get('account_number', ''),
                            'account_type': row.get('account_type', 'Individual'),
                            'account_status': row.get('account_status', 'Active'),
                            'gender': row.get('gender', ''),
                            'occupation': row.get('occupation', ''),
                            'employer_name': row.get('employer_name', ''),
                            'source_of_funds': row.get('source_of_funds', ''),
                            'is_draft': False  # Bulk imported profiles are never drafts
                        }
                        
                        # Remove None values and create the profile
                        profile_data = {k: v for k, v in data.items() if v is not None}
                        profile = KYCProfile(**profile_data)
                        profile.completion_percentage = 100  # Imported profiles are considered complete
                        profile.save()
                        
                        print(f"DEBUG: Created KYC profile with ID {profile.id}, customer_id: {profile.customer_id}")
                        
                        # Ensure workflow state exists and set to SUBMITTED
                        workflow_state = profile.workflow_state
                        workflow_state.transition_to(
                            'SUBMITTED',
                            user=request.user.username,
                            notes=f'Bulk imported on {timezone.now().strftime("%Y-%m-%d %H:%M")}'
                        )
                        
                        print(f"DEBUG: Transitioned workflow state to SUBMITTED for profile {profile.id}")
                        
                        stats['success'] += 1
                    except Exception as e:
                        stats['failed'] += 1
                        error_msg = f"Error processing row {stats['total']}: {str(e)}"
                        stats['errors'].append(error_msg)
                        print(f"DEBUG: {error_msg}")
                        logging.error(error_msg)
            
            else:  # Business import
                # Process business profiles
                for idx, row in df.iterrows():
                    stats['total'] += 1
                    try:
                        print(f"DEBUG: Processing business profile row {idx+1}: {row['business_name'] if 'business_name' in row else 'Unknown'}")
                        
                        # Create business from row data
                        data = {
                            'business_id': row.get('business_id', f"BUS-{stats['total']}"),
                            'business_name': row['business_name'],
                            'registration_number': row.get('registration_number', ''),
                            'registration_date': row.get('registration_date'),
                            'business_type': row.get('business_type', 'Corporation'),
                            'industry_sector': row.get('industry', row.get('industry_sector', 'Other')),
                            'registration_country': row.get('country_of_registration', row.get('registration_country', '')),
                            'business_address': row.get('business_address', ''),
                            'business_email': row.get('contact_email', row.get('business_email', '')),
                            'business_phone': row.get('contact_phone', row.get('business_phone', '')),
                            'annual_revenue': row.get('annual_revenue', ''),
                            'bank_name': row.get('bank_name', ''),
                            'account_number': row.get('account_number', ''),
                            'account_type': row.get('account_type', 'checking'),
                            'business_purpose': row.get('business_purpose', 'Not specified'),
                            'source_of_funds': row.get('source_of_funds', 'business_revenue'),
                            'transaction_volume': row.get('transaction_volume', 'less_than_10k'),
                            'swift_code': row.get('swift_code', ''),
                            'is_draft': False  # Bulk imported businesses are never drafts
                        }
                        
                        # Remove None values and create the business profile
                        business_data = {k: v for k, v in data.items() if v is not None}
                        business = KYCBusiness(**business_data)
                        business.save()
                        
                        print(f"DEBUG: Created KYC business with ID {business.id}, business_id: {business.business_id}")
                        
                        # Process beneficial owners if included
                        if 'beneficial_owners' in row and row['beneficial_owners']:
                            # Parse beneficial owners from a formatted string or JSON
                            try:
                                owner_data_str = str(row['beneficial_owners'])
                                print(f"DEBUG: Processing beneficial owners: {owner_data_str}")
                                
                                owners = json.loads(owner_data_str)
                                for owner in owners:
                                    business.add_beneficial_owner(owner)
                                    print(f"DEBUG: Added beneficial owner: {owner.get('full_name')}")
                            except Exception as e:
                                print(f"DEBUG: Error processing beneficial owners: {str(e)}")
                        
                        # Ensure workflow state exists and set to SUBMITTED
                        workflow_state = business.workflow_state
                        workflow_state.transition_to(
                            'SUBMITTED',
                            user=request.user.username,
                            notes=f'Bulk imported on {timezone.now().strftime("%Y-%m-%d %H:%M")}'
                        )
                        
                        print(f"DEBUG: Transitioned workflow state to SUBMITTED for business {business.id}")
                        
                        stats['success'] += 1
                    except Exception as e:
                        stats['failed'] += 1
                        error_msg = f"Error processing row {stats['total']}: {str(e)}"
                        stats['errors'].append(error_msg)
                        print(f"DEBUG: {error_msg}")
                        logging.error(error_msg)
            
            # Store import stats in session for later use
            import_results = {
                'total': stats['total'],
                'success': stats['success'],
                'failed': stats['failed'],
                'errors': stats['errors'][:10],  # Limit to first 10 errors
                'completed': True,
                'import_type': import_type
            }
            
            # Check if this is an AJAX request
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse(import_results)
            
            # Store results in session for traditional request
            request.session['import_results'] = import_results
            
            # Prepare success message
            success_message = (
                f"Import completed: {stats['success']} {import_type} profiles imported successfully, "
                f"{stats['failed']} failed."
            )
            messages.success(request, success_message)
            
            # If there were errors, add them as warning messages
            if stats['errors']:
                for error in stats['errors'][:5]:  # Show first 5 errors
                    messages.warning(request, error)
                
                if len(stats['errors']) > 5:
                    messages.warning(request, f"...and {len(stats['errors']) - 5} more errors.")
            
            # Redirect to the dashboard (or back to this page)
            return redirect('kyc_app:kyc_workflow_dashboard')
            
        except Exception as e:
            error_msg = f"Error processing Excel file: {str(e)}"
            import_results = {
                'total': 0,
                'success': 0,
                'failed': 1,
                'errors': [error_msg],
                'completed': True
            }
            
            # Check if this is an AJAX request
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse(import_results)
                
            messages.error(request, error_msg)
            print(f"DEBUG: {error_msg}")
            logging.error(error_msg)
            
            # Store error in session
            request.session['import_results'] = import_results
    
    # For AJAX requests, return a simple response for GET
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'status': 'ready'})
    
    # Check if we have results to display
    import_results = request.session.get('import_results', {})
    if import_results.get('completed'):
        # Clear the session data after retrieving it
        del request.session['import_results']
        request.session.modified = True
    
    # GET request or error in POST - show the upload form
    context = {
        'import_results': import_results if import_results.get('completed') else None
    }
    return render(request, 'bulk_import_profiles.html', context)

@login_required
def download_import_template(request):
    """
    Provides template Excel files for bulk import of KYC profiles or businesses.
    """
    import_type = request.GET.get('type', 'individual')
    
    # Create workbook with the appropriate columns
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    ws = wb.active
    
    # Add formatting
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    if import_type == 'individual':
        ws.title = "Individual Profiles"
        # Add headers for individual profiles
        headers = [
            'customer_id', 'full_name', 'date_of_birth', 'nationality', 
            'gender', 'id_document_type', 'id_document_number', 'id_issued_country',
            'id_expiry_date', 'email', 'phone_number', 'address', 'city', 'country',
            'occupation', 'employer_name', 'source_of_funds', 'account_number', 'account_type', 
            'account_status'
        ]
        # Add sample data row
        sample_data = [
            'IMP001', 'John Doe', '1980-01-01', 'USA',
            'Male', 'Passport', 'P12345678', 'USA',
            '2027-01-01', 'john@example.com', '+1234567890', '123 Main St', 'New York', 'USA',
            'Engineer', 'ABC Company', 'Salary-Formal', '1234567890', 'Individual', 'Active'
        ]
    else:  # Business template
        ws.title = "Business Profiles"
        # Add headers for business profiles
        headers = [
            'business_id', 'business_name', 'registration_number', 'registration_date',
            'business_type', 'industry_sector', 'registration_country', 
            'business_address', 'business_email', 'business_phone', 'annual_revenue',
            'business_purpose', 'transaction_volume', 'bank_name', 'account_number',
            'account_type', 'swift_code', 'beneficial_owners'
        ]
        # Add sample data row with more descriptive examples
        sample_data = [
            'BUS001', 'Acme Corporation', 'REG123456', '2010-01-01',
            'Corporation', 'Technology', 'USA',
            '456 Business Ave, Suite 100', 'contact@acme.com', '+1987654321', 
            '1m_5m',  # More than $1M annual revenue
            'Software Development and IT Services', 
            '100k_500k',  # Transaction volume
            'International Bank', '987654321', 
            'checking',  # Account type
            'INTLUS12345',  # SWIFT code
            '[{"full_name": "Jane Smith", "nationality": "USA", "id_document_type": "Passport", "id_document_number": "P987654321", "ownership_percentage": 51, "pep_status": "no"}]'
        ]
    
    # Write headers with formatting
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[chr(64 + col_idx)].width = 20  # Set column width
    
    # Write sample data
    for col_idx, value in enumerate(sample_data, 1):
        ws.cell(row=2, column=col_idx, value=value)
    
    # Add a notes row explaining the template
    ws.append([])  # Empty row
    notes_row = len(sample_data) + 2
    ws.cell(row=notes_row, column=1, value="Note: Fields in bold are required. Date format should be YYYY-MM-DD.")
    ws.merge_cells(f'A{notes_row}:E{notes_row}')
    
    # Prepare file for download
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Create response
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"kyc_{import_type}_template.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response

@login_required
def view_rejection_reason(request, profile_id):
    """
    View for displaying detailed information about a rejected KYC profile.
    Supports both individual and business profiles.
    """
    is_business = request.GET.get('is_business', '0') == '1'
    
    if is_business:
        # Handle business KYC profile
        profile = get_object_or_404(KYCBusiness, id=profile_id)
        # Get the associated workflow state
        workflow_state = KYCWorkflowState.objects.filter(
            business_kyc=profile,
            current_state='REJECTED'
        ).first()
        
        # Get test result if available
        try:
            test_result = KYCBusinessTestResult.objects.filter(business_kyc=profile).first()
        except NameError:
            # KYCBusinessTestResult model doesn't exist, handle gracefully
            test_result = None
    else:
        # Handle individual KYC profile
        profile = get_object_or_404(KYCProfile, id=profile_id)
        # Get the associated workflow state
        workflow_state = KYCWorkflowState.objects.filter(
            kyc_profile=profile,
            current_state='REJECTED'
        ).first()
        
        # Get test result if available
        test_result = KYCTestResult.objects.filter(kyc_profile=profile).first()
    
    # If we couldn't find a rejected workflow state, try to find any workflow state
    if not workflow_state:
        if is_business:
            workflow_state = KYCWorkflowState.objects.filter(business_kyc=profile).first()
        else:
            workflow_state = KYCWorkflowState.objects.filter(kyc_profile=profile).first()
    
    # Calculate date-related variables
    today = timezone.now().date()
    thirty_days_from_now = today + timedelta(days=30)
    
    context = {
        'profile': profile,
        'test_result': test_result,
        'workflow_state': workflow_state,
        'today': today,
        'thirty_days_from_now': thirty_days_from_now,
        'is_business': is_business,
    }
    
    return render(request, 'view_rejection_reason.html', context)

@login_required
@staff_member_required
def view_approval_reason(request, profile_id):
    today = datetime.now().date().strftime('%Y-%m-%d')
    thirty_days_from_now = (datetime.now().date() + timedelta(days=30)).strftime('%Y-%m-%d')
    
    is_business = request.GET.get('is_business', '0') == '1'
    
    if not is_business:
        profile = get_object_or_404(KYCProfile, id=profile_id)
        # Find the workflow state for this profile
        try:
            workflow_state = KYCWorkflowState.objects.filter(
                kyc_profile=profile, current_state='APPROVED'
            ).latest('updated_at')
        except KYCWorkflowState.DoesNotExist:
            # Fallback to any workflow state if approved one doesn't exist
            workflow_state = KYCWorkflowState.objects.filter(
                kyc_profile=profile
            ).latest('updated_at')
        
        try:
            test_result = KYCTestResult.objects.filter(kyc_profile=profile).latest('created_at')
        except KYCTestResult.DoesNotExist:
            test_result = None
            
        # Calculate next review date based on risk level if not already set
        if not hasattr(workflow_state, 'next_review_date') or not workflow_state.next_review_date:
            if test_result and test_result.risk_level:
                if test_result.risk_level == 'High':
                    # 3 months for high risk
                    months_to_add = 3
                elif test_result.risk_level == 'Medium':
                    # 6 months for medium risk
                    months_to_add = 6
                else:
                    # 12 months for low risk
                    months_to_add = 12
                
                # Calculate the date - we'll use our template filter to display it
                workflow_state.calculated_next_review_date = None  # This is just a placeholder, calculation happens in template
                
    else:
        profile = get_object_or_404(KYCBusiness, id=profile_id)
        # Find the workflow state for this profile
        try:
            workflow_state = KYCWorkflowState.objects.filter(
                business_kyc=profile, current_state='APPROVED'
            ).latest('updated_at')
        except KYCWorkflowState.DoesNotExist:
            # Fallback to any workflow state if approved one doesn't exist
            workflow_state = KYCWorkflowState.objects.filter(
                business_kyc=profile
            ).latest('updated_at')
        
        # Business profiles don't have test results in this model
        test_result = None
            
        # Calculate next review date based on default risk level for businesses
        if not hasattr(workflow_state, 'next_review_date') or not workflow_state.next_review_date:
            # Default to medium risk (6 months) for businesses
            months_to_add = 6
            workflow_state.calculated_next_review_date = None  # This is just a placeholder, calculation happens in template
    
    context = {
        'profile': profile,
        'workflow_state': workflow_state,
        'test_result': test_result,
        'today': today,
        'thirty_days_from_now': thirty_days_from_now,
        'is_business': is_business,
    }
    
    return render(request, 'view_approval_reason.html', context)

@login_required
@staff_member_required
def reopen_kyc_review(request, profile_id):
    """
    View to reopen a KYC review for a profile that has already been approved or rejected.
    This resets the workflow state back to SCREENING.
    """
    # Check if this is a business KYC profile
    if request.GET.get('type') == 'business':
        # Handle business KYC profile
        profile = get_object_or_404(KYCBusiness, id=profile_id)
        # Get the associated workflow state
        workflow_state = KYCWorkflowState.objects.filter(
            business_kyc=profile
        ).first()
        
        if workflow_state:
            # Update the workflow state to SCREENING (to be reviewed again)
            previous_state = workflow_state.current_state
            workflow_state.current_state = 'SCREENING'
            
            # Add the state transition to the history JSONField
            state_history = {
                'from_state': previous_state,
                'to_state': 'SCREENING',
                'timestamp': timezone.now().isoformat(),
                'by_user': request.user.username,
                'notes': "Profile reopened for review"
            }
            
            # Get the current history and append the new entry
            history_list = workflow_state.history if workflow_state.history else []
            history_list.append(state_history)
            workflow_state.history = history_list
            
            workflow_state.save()
            
            messages.success(request, "Business KYC profile has been reopened for review.")
    else:
        # Handle individual KYC profile
        profile = get_object_or_404(KYCProfile, id=profile_id)
        # Get the associated workflow state
        workflow_state = KYCWorkflowState.objects.filter(
            kyc_profile=profile
        ).first()
        
        if workflow_state:
            # Update the workflow state to SCREENING (to be reviewed again)
            previous_state = workflow_state.current_state
            workflow_state.current_state = 'SCREENING'
            
            # Add the state transition to the history JSONField
            state_history = {
                'from_state': previous_state,
                'to_state': 'SCREENING',
                'timestamp': timezone.now().isoformat(),
                'by_user': request.user.username,
                'notes': "Profile reopened for review"
            }
            
            # Get the current history and append the new entry
            history_list = workflow_state.history if workflow_state.history else []
            history_list.append(state_history)
            workflow_state.history = history_list
            
            workflow_state.save()
            
            messages.success(request, "KYC profile has been reopened for review.")
    
    return redirect('kyc_app:kyc_workflow_dashboard')

@login_required
def kyc_reports_list(request):
    """
    View to list all KYC reports. Provides filtering options.
    """
    # Get filter parameters
    report_type = request.GET.get('type', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    decision = request.GET.get('decision', '')
    search_query = request.GET.get('q', '')
    
    # Start with all reports
    reports = KYCReport.objects.all()
    
    # Apply filters
    if report_type:
        reports = reports.filter(report_type=report_type)
    
    if date_from:
        try:
            from_date = datetime.strptime(date_from, '%Y-%m-%d')
            reports = reports.filter(generated_at__gte=from_date)
        except ValueError:
            pass
    
    if date_to:
        try:
            to_date = datetime.strptime(date_to, '%Y-%m-%d')
            to_date = to_date.replace(hour=23, minute=59, second=59)
            reports = reports.filter(generated_at__lte=to_date)
        except ValueError:
            pass
    
    if decision:
        reports = reports.filter(decision=decision)
    
    if search_query:
        reports = reports.filter(
            Q(report_id__icontains=search_query) |
            Q(kyc_profile__full_name__icontains=search_query) |
            Q(business_kyc__business_name__icontains=search_query)
        )
    
    # Stats for the dashboard
    total_reports = reports.count()
    approved_reports = reports.filter(decision='APPROVED').count()
    rejected_reports = reports.filter(decision='REJECTED').count()
    individual_reports = reports.filter(report_type='INDIVIDUAL').count()
    business_reports = reports.filter(report_type='BUSINESS').count()
    
    context = {
        'reports': reports,
        'total_reports': total_reports,
        'approved_reports': approved_reports,
        'rejected_reports': rejected_reports,
        'individual_reports': individual_reports,
        'business_reports': business_reports,
        'filter_type': report_type,
        'filter_date_from': date_from,
        'filter_date_to': date_to,
        'filter_decision': decision,
        'search_query': search_query
    }
    
    return render(request, 'kyc_reports_list.html', context)

@login_required
def view_kyc_report(request, report_id):
    """
    View to display a single KYC report.
    """
    report = get_object_or_404(KYCReport, report_id=report_id)
    
    # Get related workflow history
    if report.kyc_profile:
        try:
            workflow_state = KYCWorkflowState.objects.get(kyc_profile=report.kyc_profile)
            # Get history from the JSONField
            json_history = workflow_state.history if workflow_state.history else []
            # Convert history entries to objects with attributes for template compatibility
            history = []
            for entry in json_history:
                if 'timestamp' in entry and 'from_state' in entry and 'to_state' in entry:
                    try:
                        history.append(type('HistoryEntry', (), {
                            'timestamp': timezone.datetime.fromisoformat(entry.get('timestamp')),
                            'from_state': entry.get('from_state'),
                            'to_state': entry.get('to_state'),
                            'user': entry.get('by_user'),
                            'notes': entry.get('notes')
                        }))
                    except (ValueError, TypeError):
                        # Skip if timestamp isn't valid
                        pass
            # Sort history by timestamp in descending order
            history.sort(key=lambda x: x.timestamp, reverse=True)
        except KYCWorkflowState.DoesNotExist:
            history = []
    elif report.business_kyc:
        try:
            workflow_state = KYCWorkflowState.objects.get(business_kyc=report.business_kyc)
            # Get history from the JSONField
            json_history = workflow_state.history if workflow_state.history else []
            # Convert history entries to objects with attributes for template compatibility
            history = []
            for entry in json_history:
                if 'timestamp' in entry and 'from_state' in entry and 'to_state' in entry:
                    try:
                        history.append(type('HistoryEntry', (), {
                            'timestamp': timezone.datetime.fromisoformat(entry.get('timestamp')),
                            'from_state': entry.get('from_state'),
                            'to_state': entry.get('to_state'),
                            'user': entry.get('by_user'),
                            'notes': entry.get('notes')
                        }))
                    except (ValueError, TypeError):
                        # Skip if timestamp isn't valid
                        pass
            # Sort history by timestamp in descending order
            history.sort(key=lambda x: x.timestamp, reverse=True)
        except KYCWorkflowState.DoesNotExist:
            history = []
    else:
        history = []
    
    context = {
        'report': report,
        'workflow_history': history
    }
    
    return render(request, 'view_kyc_report.html', context)

@login_required
def generate_kyc_report(request, profile_id, report_type):
    """
    View to generate a new KYC report based on a completed workflow.
    This can be triggered after approving or rejecting a profile.
    """
    try:
        if report_type == 'individual':
            profile = get_object_or_404(KYCProfile, id=profile_id)
            business = None
            report_type_code = 'INDIVIDUAL'
            
            # Get the workflow state
            try:
                workflow_state = KYCWorkflowState.objects.get(kyc_profile=profile)
            except KYCWorkflowState.DoesNotExist:
                messages.error(request, "No workflow state found for this profile.")
                return redirect('kyc_app:kyc_workflow_dashboard')
            
            # Get the test result
            try:
                test_result = KYCTestResult.objects.filter(kyc_profile=profile).latest('created_at')
            except KYCTestResult.DoesNotExist:
                test_result = None
                
        elif report_type == 'business':
            business = get_object_or_404(KYCBusiness, id=profile_id)
            profile = None
            report_type_code = 'BUSINESS'
            
            # Get the workflow state
            try:
                workflow_state = KYCWorkflowState.objects.get(business_kyc=business)
            except KYCWorkflowState.DoesNotExist:
                messages.error(request, "No workflow state found for this business.")
                return redirect('kyc_app:kyc_workflow_dashboard')
            
            # Get the test result
            try:
                test_result = KYCTestResult.objects.filter(business_kyc=business).latest('created_at')
            except KYCTestResult.DoesNotExist:
                test_result = None
        else:
            messages.error(request, "Invalid report type specified.")
            return redirect('kyc_app:kyc_workflow_dashboard')
        
        # Check if a report already exists and delete it if found
        existing_report = None
        if profile:
            existing_report = KYCReport.objects.filter(kyc_profile=profile).first()
        elif business:
            existing_report = KYCReport.objects.filter(business_kyc=business).first()
        
        if existing_report:
            # Delete the existing report instead of redirecting
            logger.info(f"Deleting existing report {existing_report.report_id} for {profile_id if profile else business.id}")
            existing_report.delete()
            messages.info(request, "Existing report found and deleted. Generating new report.")
        
        # Get decision and reason from workflow state
        decision = workflow_state.current_state
        if decision not in ['APPROVED', 'REJECTED']:
            messages.error(request, "Cannot generate report: Profile must be approved or rejected.")
            return redirect('kyc_app:kyc_workflow_dashboard')
        
        decision_reason = workflow_state.reviewer_notes or "No specific reason provided."
        
        # Create risk assessment text
        risk_assessment = "Risk Assessment:\n"
        if test_result:
            risk_assessment += f"Risk Level: {test_result.risk_level}\n"
            
            if test_result.sanctions_list_check:
                risk_assessment += "- Sanctions List Match: YES\n"
            else:
                risk_assessment += "- Sanctions List Match: NO\n"
                
            if test_result.politically_exposed_person:
                risk_assessment += "- Politically Exposed Person: YES\n"
            else:
                risk_assessment += "- Politically Exposed Person: NO\n"
                
            if test_result.suspicious_activity_flag:
                risk_assessment += "- Suspicious Activity Detected: YES\n"
            else:
                risk_assessment += "- Suspicious Activity Detected: NO\n"
                
            if test_result.adverse_media_check:
                risk_assessment += "- Adverse Media: YES\n"
            else:
                risk_assessment += "- Adverse Media: NO\n"
        else:
            risk_assessment += "No risk assessment was performed."
        
        # Create summary based on profile type
        if profile:
            summary = f"KYC Report for {profile.full_name} (ID: {profile.customer_id})\n"
            summary += f"Nationality: {profile.nationality}\n"
            summary += f"Date of Birth: {profile.date_of_birth}\n"
            summary += f"ID Document: {profile.id_document_type} ({profile.id_document_number})\n"
            summary += f"Decision: {decision}"
        else:
            summary = f"KYC Report for {business.business_name} (ID: {business.business_id})\n"
            summary += f"Registration Number: {business.registration_number}\n"
            summary += f"Country of Registration: {business.registration_country}\n"
            summary += f"Business Type: {business.business_type}\n"
            summary += f"Decision: {decision}"
        
        # Create new report
        report = KYCReport(
            report_type=report_type_code,
            kyc_profile=profile,
            business_kyc=business,
            summary=summary,
            risk_assessment=risk_assessment,
            decision=decision,
            decision_reason=decision_reason,
            generated_by=request.user.username,
        )
        
        # Add screening results if available
        if test_result:
            report.sanctions_check = test_result.sanctions_list_check
            report.pep_check = test_result.politically_exposed_person
            report.adverse_media_check = test_result.adverse_media_check
            
            # Check if Enhanced Due Diligence was performed
            report.edd_performed = test_result.enhanced_due_diligence_required
            report.edd_details = "Enhanced due diligence was required and performed." if test_result.enhanced_due_diligence_required else ""
        
        report.save()
        logger.info(f"KYC Report created with ID {report.report_id}")
        
        # Generate PDF report
        pdf_path = None
        try:
            pdf_path = generate_pdf_report(request, report)
            if pdf_path:
                with open(pdf_path, 'rb') as pdf_file:
                    report.pdf_report.save(f"report_{report.report_id}.pdf", ContentFile(pdf_file.read()))
                logger.info(f"PDF report saved to database for report {report.report_id}")
        except Exception as e:
            logger.error(f"Error generating PDF for report {report.report_id}: {str(e)}", exc_info=True)
            messages.warning(request, "Report created but PDF generation failed. You can try downloading it again later.")
        finally:
            # Clean up temp file if it exists
            if pdf_path and os.path.exists(pdf_path):
                try:
                    os.remove(pdf_path)
                    logger.info(f"Temporary PDF file deleted: {pdf_path}")
                except Exception as e:
                    logger.warning(f"Failed to delete temporary PDF file {pdf_path}: {str(e)}")
        
        messages.success(request, f"KYC Report generated successfully: {report.report_id}")
        return redirect('kyc_app:view_kyc_report', report_id=report.report_id)
    
    except Exception as e:
        logger.error(f"Error generating KYC report: {str(e)}", exc_info=True)
        messages.error(request, "An error occurred while generating the report. Please try again.")
        return redirect('kyc_app:kyc_workflow_dashboard')

@login_required
def download_kyc_report(request, report_id):
    """
    View to download a generated KYC report as PDF.
    """
    pdf_file = None
    try:
        # Get the report
        report = get_object_or_404(KYCReport, report_id=report_id)
        
        # Check permissions (only allow staff or the user who owns the profile)
        if not request.user.is_staff:
            if report.kyc_profile and report.kyc_profile.user != request.user:
                return HttpResponseForbidden("You don't have permission to download this report")
            if report.business_kyc and report.business_kyc.user != request.user:
                return HttpResponseForbidden("You don't have permission to download this report")
        
        # Generate the PDF
        pdf_file = generate_pdf_report(request, report)
        if not pdf_file:
            messages.error(request, "Failed to generate PDF report. Please try again later.")
            return redirect('kyc_workflow_dashboard')
            
        # Determine the subject name for the filename
        if report.kyc_profile:
            subject_name = report.kyc_profile.full_name.replace(" ", "_")
        elif report.business_kyc:
            subject_name = report.business_kyc.business_name.replace(" ", "_")
        else:
            subject_name = "kyc_subject"
            
        # Create a safe filename
        filename = f"KYC_Report_{subject_name}_{report.report_id}.pdf"
        
        # Serve the file for download
        with open(pdf_file, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
            
    except Exception as e:
        logger.error(f"Error downloading KYC report: {str(e)}", exc_info=True)
        messages.error(request, "An error occurred while downloading the report.")
        return redirect('kyc_workflow_dashboard')
    finally:
        # Clean up the temporary file
        if pdf_file and os.path.exists(pdf_file):
            try:
                os.unlink(pdf_file)
                logger.info(f"Temporary PDF file deleted: {pdf_file}")
            except Exception as e:
                logger.warning(f"Failed to delete temporary PDF file {pdf_file}: {str(e)}")

def generate_pdf_report(request, report):
    """
    Helper function to generate a PDF report using WeasyPrint.
    """
    try:
        # Get related workflow data
        if report.kyc_profile:
            try:
                workflow_state = KYCWorkflowState.objects.get(kyc_profile=report.kyc_profile)
                # Get history from the JSONField
                json_history = workflow_state.history if workflow_state.history else []
                # Convert history entries to objects with attributes for template compatibility
                history = []
                for entry in json_history:
                    if 'timestamp' in entry and 'from_state' in entry and 'to_state' in entry:
                        try:
                            history.append(type('HistoryEntry', (), {
                                'timestamp': timezone.datetime.fromisoformat(entry.get('timestamp')),
                                'from_state': entry.get('from_state'),
                                'to_state': entry.get('to_state'),
                                'user': entry.get('by_user'),
                                'notes': entry.get('notes')
                            }))
                        except (ValueError, TypeError):
                            # Skip if timestamp isn't valid
                            pass
                # Sort history by timestamp in descending order
                history.sort(key=lambda x: x.timestamp, reverse=True)
            except KYCWorkflowState.DoesNotExist:
                history = []
                logger.warning(f"No workflow state found for KYC profile {report.kyc_profile.id}")
                
            # Get test result
            try:
                test_result = KYCTestResult.objects.filter(kyc_profile=report.kyc_profile).latest('created_at')
            except KYCTestResult.DoesNotExist:
                test_result = None
                logger.info(f"No test result found for KYC profile {report.kyc_profile.id}")
                
        elif report.business_kyc:
            try:
                workflow_state = KYCWorkflowState.objects.get(business_kyc=report.business_kyc)
                # Get history from the JSONField
                json_history = workflow_state.history if workflow_state.history else []
                # Convert history entries to objects with attributes for template compatibility
                history = []
                for entry in json_history:
                    if 'timestamp' in entry and 'from_state' in entry and 'to_state' in entry:
                        try:
                            history.append(type('HistoryEntry', (), {
                                'timestamp': timezone.datetime.fromisoformat(entry.get('timestamp')),
                                'from_state': entry.get('from_state'),
                                'to_state': entry.get('to_state'),
                                'user': entry.get('by_user'),
                                'notes': entry.get('notes')
                            }))
                        except (ValueError, TypeError):
                            # Skip if timestamp isn't valid
                            pass
                # Sort history by timestamp in descending order
                history.sort(key=lambda x: x.timestamp, reverse=True)
            except KYCWorkflowState.DoesNotExist:
                history = []
                logger.warning(f"No workflow state found for Business KYC {report.business_kyc.id}")
                
            # For business profiles, we don't have test results since KYCTestResult doesn't have business_kyc field
            test_result = None
            logger.info(f"No test result available for Business KYC {report.business_kyc.id}")
        else:
            history = []
            test_result = None
            logger.warning("Report has no associated KYC profile or Business KYC")
        
        # Get logo URL if available
        logo_url = getattr(settings, 'COMPANY_LOGO_URL', None)
        base_url = request.build_absolute_uri('/')[:-1]
        
        # Convert relative logo URL to absolute if needed
        if logo_url and not logo_url.startswith(('http://', 'https://')):
            logo_url = f"{base_url}{logo_url}"
        
        # Get decision date from workflow history or use report generation date
        decision_date = None
        if history:
            for entry in history:
                if entry.to_state in ['APPROVED', 'REJECTED']:
                    decision_date = entry.timestamp
                    break
        if not decision_date:
            decision_date = report.generated_at
        
        # Add reviewer information
        reviewer = report.generated_by
        if hasattr(workflow_state, 'reviewer_username') and workflow_state.reviewer_username:
            reviewer = workflow_state.reviewer_username
            
        # Render the HTML template
        html_string = render_to_string(
            'kyc_report_pdf.html',
            {
                'report': report,
                'workflow_history': history,
                'test_result': test_result,
                'base_url': base_url,
                'logo_url': logo_url,
                'decision_date': decision_date,
                'reviewer': reviewer
            }
        )
        
        # Create a temporary file
        output_file = os.path.join(settings.MEDIA_ROOT, f"tmp_report_{uuid.uuid4()}.pdf")
        
        # Define CSS styles for better PDF formatting
        css_string = '''
            @page {
                margin: 2cm;
                @top-center {
                    content: "KYC Report";
                    font-size: 9pt;
                }
                @bottom-right {
                    content: "Page " counter(page) " of " counter(pages);
                    font-size: 9pt;
                }
            }
            body {
                font-family: Arial, sans-serif;
                font-size: 11pt;
                line-height: 1.5;
            }
            h1, h2, h3 {
                margin-top: 20px;
                margin-bottom: 10px;
                color: #333;
            }
            table {
                width: 100%;
                border-collapse: collapse;
                margin-bottom: 15px;
            }
            th, td {
                border: 1px solid #ddd;
                padding: 8px;
                text-align: left;
            }
            th {
                background-color: #f2f2f2;
            }
            .header {
                margin-bottom: 30px;
            }
            .footer {
                margin-top: 30px;
                border-top: 1px solid #eee;
                padding-top: 10px;
                font-size: 9pt;
                color: #666;
            }
            .logo {
                max-width: 200px;
                max-height: 60px;
            }
        '''
        
        # Generate PDF
        logger.info(f"Generating PDF report for {report.id}")
        
        # Use the correct WeasyPrint approach for handling external resources
        html = HTML(string=html_string, base_url=base_url)
        pdf = html.write_pdf(stylesheets=[CSS(string=css_string)])
        
        # Save the PDF to a file
        with open(output_file, 'wb') as f:
            f.write(pdf)
        
        logger.info(f"PDF report successfully generated at {output_file}")
        return output_file
    except Exception as e:
        logger.error(f"PDF generation error: {str(e)}", exc_info=True)
        return None

@login_required
def reports_dashboard(request):
    """
    View for the reports generation dashboard showing profiles ready for report generation.
    """
    # Get individual profiles with approved or rejected status
    approved_individuals = KYCProfile.objects.filter(
        workflow_state__current_state='APPROVED',
    ).select_related('workflow_state')
    
    rejected_individuals = KYCProfile.objects.filter(
        workflow_state__current_state='REJECTED',
    ).select_related('workflow_state')
    
    # Get business profiles with approved or rejected status
    approved_businesses = KYCBusiness.objects.filter(
        workflow_state__current_state='APPROVED',
    ).select_related('workflow_state')
    
    rejected_businesses = KYCBusiness.objects.filter(
        workflow_state__current_state='REJECTED',
    ).select_related('workflow_state')
    
    # Add risk level information to profiles
    individuals = list(approved_individuals) + list(rejected_individuals)
    businesses = list(approved_businesses) + list(rejected_businesses)
    
    # Get risk levels for individual profiles
    for profile in individuals:
        try:
            test_result = KYCTestResult.objects.filter(kyc_profile=profile).latest('created_at')
            profile.risk_level = test_result.risk_level
        except KYCTestResult.DoesNotExist:
            profile.risk_level = None
    
    # Get risk levels for business profiles
    # Note: It appears KYCTestResult doesn't have a business_kyc field
    # So we'll set risk_level to None for all businesses
    for business in businesses:
        business.risk_level = None
        # If you want to implement business risk levels, you would need to
        # modify the KYCTestResult model to include a business_kyc field
    
    # Get recently generated reports
    recent_reports = KYCReport.objects.all().order_by('-generated_at')[:10]
    
    context = {
        'individuals': individuals,
        'businesses': businesses,
        'approved_individuals': approved_individuals.count(),
        'rejected_individuals': rejected_individuals.count(),
        'approved_businesses': approved_businesses.count(),
        'rejected_businesses': rejected_businesses.count(),
        'recent_reports': recent_reports,
    }
    
    return render(request, 'generate_reports.html', context)

@login_required
def batch_generate_reports(request):
    """
    View to handle batch generation of reports for multiple profiles.
    """
    if request.method != 'POST':
        messages.error(request, "Invalid request method.")
        return redirect('kyc_app:reports_dashboard')
    
    profile_type = request.POST.get('profile_type', 'individual')
    status = request.POST.get('status', 'all')
    
    # Set up filter based on status
    if status == 'APPROVED':
        status_filter = {'workflow_state__current_state': 'APPROVED'}
    elif status == 'REJECTED':
        status_filter = {'workflow_state__current_state': 'REJECTED'}
    else:  # All completed profiles
        status_filter = {'workflow_state__current_state__in': ['APPROVED', 'REJECTED']}
    
    # Stats for tracking processing
    report_stats = {
        'total': 0,
        'success': 0,
        'failed': 0,
        'errors': [],
        'replaced': 0
    }
    
    try:
        if profile_type == 'individual':
            # Get individual profiles based on filter
            profiles = KYCProfile.objects.filter(**status_filter).select_related('workflow_state')
            
            for profile in profiles:
                report_stats['total'] += 1
                try:
                    # Check if a report already exists for this profile and delete it
                    existing_report = KYCReport.objects.filter(kyc_profile=profile).first()
                    if existing_report:
                        logger.info(f"Deleting existing report {existing_report.report_id} for profile {profile.id}")
                        existing_report.delete()
                        report_stats['replaced'] += 1
                    
                    # Get the workflow state
                    workflow_state = profile.workflow_state
                    
                    # Create risk assessment text
                    risk_assessment = "Risk Assessment:\n"
                    try:
                        test_result = KYCTestResult.objects.filter(kyc_profile=profile).latest('created_at')
                        risk_assessment += f"Risk Level: {test_result.risk_level}\n"
                        
                        risk_assessment += f"Sanctions List Match: {'YES' if test_result.sanctions_list_check else 'NO'}\n"
                        risk_assessment += f"Politically Exposed Person: {'YES' if test_result.politically_exposed_person else 'NO'}\n"
                        risk_assessment += f"Suspicious Activity Detected: {'YES' if test_result.suspicious_activity_flag else 'NO'}\n"
                        risk_assessment += f"Adverse Media: {'YES' if test_result.adverse_media_check else 'NO'}\n"
                    except KYCTestResult.DoesNotExist:
                        test_result = None
                        risk_assessment += "No risk assessment was performed."
                    
                    # Create summary
                    summary = f"KYC Report for {profile.full_name} (ID: {profile.customer_id})\n"
                    summary += f"Nationality: {profile.nationality}\n"
                    summary += f"Date of Birth: {profile.date_of_birth}\n"
                    summary += f"ID Document: {profile.id_document_type} ({profile.id_document_number})\n"
                    summary += f"Decision: {workflow_state.current_state}"
                    
                    # Create report
                    report = KYCReport(
                        report_type='INDIVIDUAL',
                        kyc_profile=profile,
                        summary=summary,
                        risk_assessment=risk_assessment,
                        decision=workflow_state.current_state,
                        decision_reason=workflow_state.reviewer_notes or "No specific reason provided.",
                        generated_by=request.user.username,
                    )
                    
                    # Add screening results if available
                    if test_result:
                        report.sanctions_check = test_result.sanctions_list_check
                        report.pep_check = test_result.politically_exposed_person
                        report.adverse_media_check = test_result.adverse_media_check
                        report.edd_performed = test_result.enhanced_due_diligence_required
                        report.edd_details = "Enhanced due diligence was required and performed." if test_result.enhanced_due_diligence_required else ""
                    
                    report.save()
                    report_stats['success'] += 1
                    
                    # Generate PDF report
                    try:
                        pdf_path = generate_pdf_report(request, report)
                        if pdf_path:
                            with open(pdf_path, 'rb') as pdf_file:
                                report.pdf_report.save(f"report_{report.report_id}.pdf", ContentFile(pdf_file.read()))
                            
                            # Clean up temp file
                            if os.path.exists(pdf_path):
                                os.remove(pdf_path)
                    except Exception as e:
                        logger.error(f"PDF generation error for profile {profile.id}: {str(e)}", exc_info=True)
                
                except Exception as e:
                    report_stats['failed'] += 1
                    error_msg = f"Error processing profile {profile.id}: {str(e)}"
                    report_stats['errors'].append(error_msg)
                    logger.error(error_msg, exc_info=True)
        
        else:  # Business profiles
            # Get business profiles based on filter
            businesses = KYCBusiness.objects.filter(**status_filter).select_related('workflow_state')
            
            for business in businesses:
                report_stats['total'] += 1
                try:
                    # Check if a report already exists for this business and delete it
                    existing_report = KYCReport.objects.filter(business_kyc=business).first()
                    if existing_report:
                        logger.info(f"Deleting existing report {existing_report.report_id} for business {business.id}")
                        existing_report.delete()
                        report_stats['replaced'] += 1
                    
                    # Get the workflow state
                    workflow_state = business.workflow_state
                    
                    # Create risk assessment text - no test results for businesses
                    risk_assessment = "Risk Assessment:\n"
                    # Since KYCTestResult doesn't have business_kyc field, we can't get test results
                    risk_assessment += "No risk assessment data available for business profiles."
                    test_result = None
                    
                    # Create summary
                    summary = f"KYC Report for {business.business_name} (ID: {business.business_id})\n"
                    summary += f"Registration Number: {business.registration_number}\n"
                    summary += f"Country of Registration: {business.registration_country}\n"
                    summary += f"Business Type: {business.business_type}\n"
                    summary += f"Decision: {workflow_state.current_state}"
                    
                    # Create report
                    report = KYCReport(
                        report_type='BUSINESS',
                        business_kyc=business,
                        summary=summary,
                        risk_assessment=risk_assessment,
                        decision=workflow_state.current_state,
                        decision_reason=workflow_state.reviewer_notes or "No specific reason provided.",
                        generated_by=request.user.username,
                    )
                    
                    # We don't have test results for businesses
                    report.sanctions_check = False
                    report.pep_check = False
                    report.adverse_media_check = False
                    report.edd_performed = False
                    report.edd_details = ""
                    
                    report.save()
                    report_stats['success'] += 1
                    
                    # Generate PDF report
                    try:
                        pdf_path = generate_pdf_report(request, report)
                        if pdf_path:
                            with open(pdf_path, 'rb') as pdf_file:
                                report.pdf_report.save(f"report_{report.report_id}.pdf", ContentFile(pdf_file.read()))
                            
                            # Clean up temp file
                            if os.path.exists(pdf_path):
                                os.remove(pdf_path)
                    except Exception as e:
                        logger.error(f"PDF generation error for business {business.id}: {str(e)}", exc_info=True)
                
                except Exception as e:
                    report_stats['failed'] += 1
                    error_msg = f"Error processing business {business.id}: {str(e)}"
                    report_stats['errors'].append(error_msg)
                    logger.error(error_msg, exc_info=True)
        
        # Prepare success message
        if report_stats['success'] > 0:
            replaced_msg = f" ({report_stats['replaced']} existing reports replaced)" if report_stats['replaced'] > 0 else ""
            messages.success(
                request, 
                f"Successfully generated {report_stats['success']} reports for {profile_type} profiles{replaced_msg}."
            )
        else:
            messages.warning(
                request, 
                f"No reports were generated for {profile_type} profiles."
            )
        
        # Show errors if any
        if report_stats['failed'] > 0:
            for error in report_stats['errors'][:3]:  # Show only first 3 errors
                messages.error(request, error)
            
            if len(report_stats['errors']) > 3:
                messages.error(request, f"...and {len(report_stats['errors']) - 3} more errors.")
    
    except Exception as e:
        messages.error(request, f"An error occurred during batch report generation: {str(e)}")
        logger.error(f"Batch report generation error: {str(e)}", exc_info=True)
    
    return redirect('kyc_app:reports_dashboard')

@login_required
@user_passes_test(lambda u: u.is_staff)
def document_verification_dashboard(request):
    """
    Display a dashboard for document verification with filtering capabilities.
    """
    # Get filter parameters from request
    document_type = request.GET.get('document_type', '')
    status = request.GET.get('status', '')
    search = request.GET.get('search', '')
    
    # Base queryset - get all documents that need verification
    documents = Document.objects.select_related('profile').all()
    
    # Apply filters
    if document_type:
        documents = documents.filter(document_type=document_type)
    
    if status:
        documents = documents.filter(status=status)
    
    if search:
        documents = documents.filter(
            Q(profile__customer_id__icontains=search) | 
            Q(profile__full_name__icontains=search) |
            Q(profile__email__icontains=search)
        )
    
    # Default sort by upload date (newest first)
    documents = documents.order_by('-upload_date')
    
    # Calculate dashboard metrics
    today = timezone.now().date()
    pending_count = documents.filter(status='PENDING').count()
    verified_today = documents.filter(status='VERIFIED', verification_date__date=today).count()
    rejected_today = documents.filter(status='REJECTED', verification_date__date=today).count()
    
    # High priority documents - get pending docs from profiles with high risk level from KYCTestResult
    # First get the high risk profile IDs
    high_risk_profile_ids = KYCTestResult.objects.filter(
        risk_level='High'
    ).values_list('kyc_profile_id', flat=True).distinct()
    
    # Then filter documents based on those profile IDs
    high_priority_count = documents.filter(
        status='PENDING',
        profile_id__in=high_risk_profile_ids
    ).count()
    
    # Pagination
    paginator = Paginator(documents, 15)  # Show 15 documents per page
    page = request.GET.get('page')
    try:
        documents = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page
        documents = paginator.page(1)
    except EmptyPage:
        # If page is out of range, deliver last page of results
        documents = paginator.page(paginator.num_pages)
    
    context = {
        'documents': documents,
        'pending_count': pending_count,
        'verified_today': verified_today,
        'rejected_today': rejected_today,
        'high_priority_count': high_priority_count,
        'document_type': document_type,
        'status': status,
        'search': search,
    }
    
    return render(request, 'document_verification_dashboard.html', context)

@login_required
def verify_document(request, document_id):
    """
    View for verifying or rejecting a specific document.
    This allows staff to review the document and update its status.
    """
    # Force session save to prevent timeouts
    request.session.modified = True
    
    try:
        document = get_object_or_404(Document, id=document_id)
    except Document.DoesNotExist:
        messages.error(request, "Document not found.")
        return redirect('kyc_app:document_verification_dashboard')
    
    # Check if this is an AJAX request
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    
    # If this is a POST request, process the form data
    if request.method == 'POST':
        action = request.POST.get('action')
        notes = request.POST.get('notes', '')
        rejection_reason = request.POST.get('rejection_reason', '')
        
        # Check if the no_expiry checkbox is checked
        no_expiry = request.POST.get('no_expiry') == 'on'
        
        # Update document details
        document_number = request.POST.get('document_number')
        if document_number:
            document.document_number = document_number
        else:
            # If document number is required but not provided, show an error
            messages.error(request, "Document number is required.")
            context = {
                'document': document,
                'next': request.POST.get('next', ''),
                'profile': document.profile,
            }
            return render(request, 'verify_document.html', context)
            
        issue_date = request.POST.get('issue_date')
        if issue_date:
            try:
                document.issue_date = datetime.strptime(issue_date, '%Y-%m-%d').date()
            except (ValueError, TypeError):
                pass
                
        # Handle expiry date based on the no_expiry checkbox
        if no_expiry:
            document.expiry_date = None
        else:
            expiry_date = request.POST.get('expiry_date')
            if expiry_date:
                try:
                    document.expiry_date = datetime.strptime(expiry_date, '%Y-%m-%d').date()
                except (ValueError, TypeError):
                    pass
                
        document.issuing_authority = request.POST.get('issuing_authority', document.issuing_authority)
        
        # If issuing country is not provided, use the customer's nationality
        issuing_country = request.POST.get('issuing_country')
        if issuing_country:
            document.issuing_country = issuing_country
        elif not document.issuing_country and document.profile.nationality:
            document.issuing_country = document.profile.nationality
            
        # Save document details first
        document.save()
        
        # Force session save again after primary operations
        request.session.save()
        
        try:
            if action == 'verify':
                document.verify(request.user.username, notes)
                messages.success(request, f"Document has been verified successfully.")
            elif action == 'reject':
                if not rejection_reason:
                    messages.error(request, "Please provide a reason for rejection.")
                    context = {
                        'document': document,
                        'next': request.POST.get('next', ''),
                        'profile': document.profile,
                    }
                    return render(request, 'verify_document.html', context)
                
                document.reject(request.user.username, rejection_reason)
                messages.success(request, f"Document has been rejected.")
                
            # Force session save after verification actions
            request.session.save()
                
        except Exception as e:
            # Log the error
            print(f"Error in document verification: {str(e)}")
            messages.error(request, f"An error occurred: {str(e)}")
            context = {
                'document': document,
                'next': request.POST.get('next', ''),
                'profile': document.profile,
            }
            return render(request, 'verify_document.html', context)
        
        # For AJAX requests, return a JSON response
        if is_ajax:
            from django.http import JsonResponse
            return JsonResponse({
                'status': 'success',
                'message': f"Document has been {'verified' if action == 'verify' else 'rejected'} successfully.",
                'redirect_url': request.POST.get('next', '') or reverse('kyc_app:document_verification_dashboard')
            })
        
        # Redirect based on next parameter or back to dashboard
        next_url = request.POST.get('next', '')
        if next_url:
            return redirect(next_url)
        return redirect('kyc_app:document_verification_dashboard')
    
    # For GET requests, pre-fill document details if they're empty
    # Auto-fill document number from profile if available
    if not document.document_number and document.profile.id_document_number:
        document.document_number = document.profile.id_document_number
    
    # Auto-fill issue date from profile's issued date if available and not already set
    if not document.issue_date and hasattr(document.profile, 'id_document_issue_date') and document.profile.id_document_issue_date:
        document.issue_date = document.profile.id_document_issue_date
    
    # Auto-fill expiry date from profile if available and not already set
    if not document.expiry_date and hasattr(document.profile, 'id_expiry_date') and document.profile.id_expiry_date:
        document.expiry_date = document.profile.id_expiry_date
    
    # Auto-fill issuing country from nationality if available
    if not document.issuing_country and document.profile.nationality:
        document.issuing_country = document.profile.nationality
    
    # Prepare context for the template
    context = {
        'document': document,
        'next': request.GET.get('next', ''),
        # Get the profile for additional context
        'profile': document.profile,
    }
    
    # Force session save before rendering template
    request.session.save()
    
    return render(request, 'verify_document.html', context)

@login_required
def browse_documents(request):
    """
    View to browse documents organized by customer/business folders.
    Displays a list of customers and businesses with their document folders.
    """
    # Get query parameters
    view_type = request.GET.get('type', 'all')  # all, individuals, businesses
    search_term = request.GET.get('search', '')
    
    # Initialize context data
    customers = []
    businesses = []
    
    # Get individual profiles with documents
    if view_type in ['all', 'individuals']:
        profiles_query = KYCProfile.objects.all()
        
        # Apply search filter if provided
        if search_term:
            profiles_query = profiles_query.filter(
                Q(customer_id__icontains=search_term) | 
                Q(full_name__icontains=search_term)
            )
            
        # Get profiles and count their documents
        for profile in profiles_query:
            # Count documents by type
            doc_counts = {}
            documents = Document.objects.filter(profile=profile)
            
            # Calculate total documents
            total_docs = documents.count()
            
            # Only include profiles with documents or document files
            has_id_document = bool(profile.id_document_file)
            if total_docs > 0 or has_id_document:
                # Count documents by type
                for doc_type, _ in Document.DOCUMENT_TYPE_CHOICES:
                    doc_counts[doc_type] = documents.filter(document_type=doc_type).count()
                
                customers.append({
                    'id': profile.id,
                    'customer_id': profile.customer_id,
                    'name': profile.full_name,
                    'document_counts': doc_counts,
                    'total_documents': total_docs + (1 if has_id_document else 0),
                    'has_id_document': has_id_document,
                    'folder_path': profile.create_document_folders()
                })
    
    # Get businesses with documents
    if view_type in ['all', 'businesses']:
        businesses_query = KYCBusiness.objects.all()
        
        # Apply search filter if provided
        if search_term:
            businesses_query = businesses_query.filter(
                Q(business_id__icontains=search_term) | 
                Q(business_name__icontains=search_term)
            )
        
        # Get businesses with documents
        for business in businesses_query:
            has_registration_doc = bool(business.registration_document)
            has_tax_doc = bool(business.tax_document)
            total_docs = (1 if has_registration_doc else 0) + (1 if has_tax_doc else 0)
            
            # Only include businesses with documents
            if total_docs > 0:
                businesses.append({
                    'id': business.id,
                    'business_id': business.business_id,
                    'name': business.business_name,
                    'has_registration_doc': has_registration_doc,
                    'has_tax_doc': has_tax_doc,
                    'total_documents': total_docs,
                    'folder_path': business.create_document_folders()
                })
    
    # Debug counts to help troubleshoot
    print(f"DEBUG: Document Browser - Found {len(customers)} customers with documents")
    for customer in customers:
        print(f"DEBUG: Customer {customer['name']} has {customer['total_documents']} documents")
    
    context = {
        'customers': customers,
        'businesses': businesses,
        'view_type': view_type,
        'search_term': search_term,
        'total_customers': len(customers),
        'total_businesses': len(businesses)
    }
    
    return render(request, 'browse_documents.html', context)

def get_document_type_from_id_type(id_type):
    """
    Maps profile/business ID document type to Document model document type.
    
    Args:
        id_type: The ID document type from KYCProfile or KYCBusiness
        
    Returns:
        String: Matching document type from Document.DOCUMENT_TYPE_CHOICES
    """
    id_type_mapping = {
        'Passport': 'PASSPORT',
        'National ID': 'ID_CARD',
        'Driver License': 'DRIVERS_LICENSE',
        'Business Registration': 'REGISTRATION_CERT',
        'Tax Certificate': 'TAX_CERT',
        'Other': 'ID_CARD'
    }
    return id_type_mapping.get(id_type, 'ID_CARD')


def import_id_document(profile=None, business=None):
    """
    Imports ID document from a KYCProfile or KYCBusiness into the Document model.
    
    Args:
        profile: KYCProfile instance (for individuals)
        business: KYCBusiness instance (for businesses)
    
    Returns:
        bool: Whether a document was successfully imported
    """
    try:
        from django.core.files.base import ContentFile
        import os
        
        if profile and profile.id_document_file:
            # Check if an ID document already exists
            doc_type = get_document_type_from_id_type(profile.id_document_type)
            existing_id_doc = Document.objects.filter(
                profile=profile,
                document_type=doc_type
            ).exists()
            
            if not existing_id_doc:
                # Create a new Document record for the ID document
                document = Document(
                    profile=profile,
                    document_type=doc_type,
                    status='PENDING',
                    verification_notes=f"Imported from profile registration ({profile.id_document_type})"
                )
                
                # Get the source file path and name
                source_path = profile.id_document_file.path
                source_name = os.path.basename(profile.id_document_file.name)
                
                # Save the document first to generate ID
                document.save()
                
                # Now assign the file content
                with open(source_path, 'rb') as f:
                    document.document_file.save(source_name, ContentFile(f.read()), save=True)
                
                return True
                
        elif business and (business.registration_document or business.tax_document):
            # Handle business documents
            imported = False
            
            # Import registration document if exists
            if business.registration_document:
                doc_type = 'REGISTRATION_CERT'
                existing_doc = Document.objects.filter(
                    business=business,
                    document_type=doc_type
                ).exists()
                
                if not existing_doc:
                    document = Document(
                        business=business,
                        document_type=doc_type,
                        status='PENDING',
                        verification_notes="Imported from business registration"
                    )
                    
                    # Get the source file path and name
                    source_path = business.registration_document.path
                    source_name = os.path.basename(business.registration_document.name)
                    
                    # Save the document first to generate ID
                    document.save()
                    
                    # Now assign the file content
                    with open(source_path, 'rb') as f:
                        document.document_file.save(source_name, ContentFile(f.read()), save=True)
                    
                    imported = True
            
            # Import tax document if exists
            if business.tax_document:
                doc_type = 'TAX_CERT'
                existing_doc = Document.objects.filter(
                    business=business,
                    document_type=doc_type
                ).exists()
                
                if not existing_doc:
                    document = Document(
                        business=business,
                        document_type=doc_type,
                        status='PENDING',
                        verification_notes="Imported from business registration"
                    )
                    
                    # Get the source file path and name
                    source_path = business.tax_document.path
                    source_name = os.path.basename(business.tax_document.name)
                    
                    # Save the document first to generate ID
                    document.save()
                    
                    # Now assign the file content
                    with open(source_path, 'rb') as f:
                        document.document_file.save(source_name, ContentFile(f.read()), save=True)
                    
                    imported = True
            
            return imported
    
    except Exception as e:
        print(f"Error importing ID document: {str(e)}")
        return False
    
    return False

@login_required
def customer_documents(request, customer_id):
    """
    View for customer document management.
    Displays documents for a specific customer and allows uploading new documents.
    """
    try:
        profile = KYCProfile.objects.get(customer_id=customer_id)
    except KYCProfile.DoesNotExist:
        messages.error(request, "Customer profile not found.")
        return redirect('kyc_app:browse_documents')
    
    # Check if the profile has an ID document file that hasn't been imported
    # into the Document model yet
    if profile.id_document_file:
        imported = import_id_document(profile=profile)
        if imported:
            messages.info(request, "The ID document uploaded during registration has been imported.")
    
    # Get all documents associated with this profile
    documents = Document.objects.filter(profile=profile)
    
    # Prepare context with profile info
    context = {
        'profile': profile,
        'customer': profile,  # For backwards compatibility with template
        'documents': documents,
        'has_id_document': bool(profile.id_document_file),
        'document_types': dict(Document.DOCUMENT_TYPE_CHOICES),
    }
    
    # Log document count for debugging
    print(f"DEBUG: Found {documents.count()} documents for customer {profile.full_name}")
    for doc in documents:
        print(f"DEBUG: Document {doc.id}: {doc.document_type} - {doc.document_file}")
    
    # Handle document upload form submission
    if request.method == 'POST' and 'document_file' in request.FILES:
        try:
            # Call the upload function
            success = upload_customer_document(request, customer_id)
            if success:
                messages.success(request, "Document uploaded successfully.")
                return redirect('kyc_app:customer_documents', customer_id=customer_id)
        except Exception as e:
            messages.error(request, f"Error uploading document: {str(e)}")
    
    return render(request, 'customer_documents.html', context)

@login_required
@user_passes_test(lambda u: u.is_staff)
def business_documents(request, business_id):
    """
    View to display all documents for a specific business.
    """
    try:
        # Get the business
        business = get_object_or_404(KYCBusiness, business_id=business_id)
        
        # Check for registration and tax documents
        has_registration_doc = bool(business.registration_document)
        has_tax_doc = bool(business.tax_document)
        
        context = {
            'business': business,
            'has_registration_doc': has_registration_doc,
            'has_tax_doc': has_tax_doc,
            'document_folder': business.create_document_folders()
        }
        
        # Handle document upload if form submitted
        if request.method == 'POST' and request.FILES:
            try:
                # Call the upload function
                success = upload_business_document(request, business_id)
                if success:
                    messages.success(request, "Business document uploaded successfully.")
                    return redirect('kyc_app:business_documents', business_id=business_id)
            except Exception as e:
                messages.error(request, f"Error uploading document: {str(e)}")
        
        return render(request, 'business_documents.html', context)
    
    except Exception as e:
        messages.error(request, f"Error loading business documents: {str(e)}")
        return redirect('kyc_app:browse_documents')

@login_required
def upload_customer_document(request, customer_id):
    """
    View to handle uploading documents for a specific customer.
    """
    try:
        # Get the profile
        profile = get_object_or_404(KYCProfile, customer_id=customer_id)
        
        if request.method == 'POST':
            # Get form data
            document_type = request.POST.get('document_type')
            document_file = request.FILES.get('document_file')
            description = request.POST.get('description', '')
            
            if not document_file:
                messages.error(request, "No file was uploaded. Please select a file.")
                return False
            
            # Map the form document types to model document types
            document_type_mapping = {
                'id_document': 'PASSPORT',  # Default to passport
                'national_id': 'ID_CARD',
                'drivers_license': 'DRIVERS_LICENSE',
                'residence_permit': 'RESIDENCE_PERMIT',
                'proof_of_address': 'UTILITY_BILL',
                'proof_of_income': 'BANK_STATEMENT',
                'bank_statement': 'BANK_STATEMENT',
                'other': 'ID_CARD',  # Default fallback
            }
            
            # Get the mapped document type
            mapped_doc_type = document_type_mapping.get(document_type, 'ID_CARD')
            
            # Check if this is an ID document upload and update the profile if needed
            if document_type in ['id_document', 'national_id', 'drivers_license']:
                # Map Document model types back to KYCProfile id_document_type
                profile_type_mapping = {
                    'PASSPORT': 'Passport',
                    'ID_CARD': 'National ID',
                    'DRIVERS_LICENSE': 'Driver License'
                }
                
                # If profile doesn't have an ID document, update it
                if not profile.id_document_file:
                    profile.id_document_type = profile_type_mapping.get(mapped_doc_type, 'Other')
                    profile.id_document_file = document_file
                    profile.save(update_fields=['id_document_type', 'id_document_file'])
                    messages.info(request, "Your profile ID document has been updated.")
            
            # Create new document
            document = Document(
                profile=profile,
                document_type=mapped_doc_type,
                document_file=document_file,
                status='PENDING',
                verification_notes=description if description else None
            )
            
            # Save the document
            document.save()
            
            messages.success(request, "Document uploaded successfully and pending verification.")
            return True
            
        # If not POST, return False
        return False
        
    except Exception as e:
        print(f"ERROR in upload_customer_document: {str(e)}")
        messages.error(request, f"Error uploading document: {str(e)}")
        return False

@login_required
def upload_business_document(request, business_id):
    """
    View to handle uploading documents for a specific business.
    """
    try:
        # Get the business
        business = get_object_or_404(KYCBusiness, business_id=business_id)
        
        if request.method == 'POST':
            # Get form data
            document_type = request.POST.get('document_type')
            document_file = request.FILES.get('document_file')
            description = request.POST.get('description', '')
            
            if not document_file:
                messages.error(request, "No file was uploaded. Please select a file.")
                return False
            
            # Handle different document types
            if document_type == 'registration_certificate':
                business.registration_document = document_file
                business.save()
                messages.success(request, "Registration document uploaded successfully.")
                return True
            elif document_type == 'tax_certificate':
                business.tax_document = document_file
                business.save()
                messages.success(request, "Tax document uploaded successfully.")
                return True
            else:
                # For other document types, we might need to create a separate model
                # This is a simplified implementation
                messages.info(request, f"Document of type '{document_type}' uploaded. Note: Currently only registration and tax documents are supported.")
                return True
        
        # If not POST, return False
        return False
        
    except Exception as e:
        print(f"ERROR in upload_business_document: {str(e)}")
        messages.error(request, f"Error uploading document: {str(e)}")
        return False

@login_required
def view_document(request, document_id):
    """
    View for displaying a document with status badges and detailed information.
    This allows users to see document details and verification status.
    """
    try:
        document = get_object_or_404(Document, id=document_id)
        
        # If the user is not staff, check if the document belongs to them
        if not request.user.is_staff:
            # Check if the user is associated with the document's profile
            # This would need to be modified based on your user-profile relationship
            if not hasattr(document.profile, 'user') or document.profile.user != request.user:
                messages.error(request, "You don't have permission to view this document.")
                return redirect('dashboard')
        
        # Get the customer information for display
        customer = None
        if hasattr(document, 'profile') and document.profile is not None:
            customer = document.profile
        elif hasattr(document, 'business') and document.business is not None:
            customer = document.business
        
        context = {
            'document': document,
            'customer': customer,
        }
        
        return render(request, 'view_document.html', context)
        
    except Document.DoesNotExist:
        messages.error(request, "Document not found.")
        return redirect('kyc_app:browse_documents')
    except Exception as e:
        messages.error(request, f"Error viewing document: {str(e)}")
        return redirect('kyc_app:browse_documents')

# Helper function to update beneficial owners for a business
def update_beneficial_owners(business, owners_data):
    """
    Updates the beneficial owners for a business using the JSON field
    
    Args:
        business: KYCBusiness instance
        owners_data: List of dictionaries with beneficial owner data
    """
    try:
        # Initialize beneficial owners list if needed
        if not business.beneficial_owners:
            business.beneficial_owners = []
        else:
            # Clear existing beneficial owners to replace with new data
            business.beneficial_owners = []
        
        # Add each beneficial owner to the JSON field
        for owner_data in owners_data:
            business.add_beneficial_owner(owner_data)
            
        return True
    except Exception as e:
        print(f"Error updating beneficial owners: {str(e)}")
        return False

# Helper function to create workflow state
def create_workflow_state(entity_id, entity_type, state='DRAFT'):
    """
    Creates a new workflow state for a profile or business
    
    Args:
        entity_id: ID of the entity (profile or business)
        entity_type: Type of entity ('profile' or 'business')
        state: Initial state ('DRAFT', 'PENDING', etc.)
        
    Returns:
        KYCWorkflowState instance
    """
    try:
        if entity_type == 'profile':
            profile = KYCProfile.objects.get(id=entity_id)
            workflow_state = KYCWorkflowState(
                kyc_profile=profile,
                current_state=state
            )
        elif entity_type == 'business':
            business = KYCBusiness.objects.get(id=entity_id)
            workflow_state = KYCWorkflowState(
                business_kyc=business,
                current_state=state
            )
        else:
            return None
            
        workflow_state.save()
        return workflow_state
    except Exception as e:
        print(f"Error creating workflow state: {str(e)}")
        return None

@login_required
def debug_field_lengths(request):
    """
    Diagnostic view to help identify field length issues in the KYCBusiness model.
    """
    from django.db import models
    
    # Get all fields from KYCBusiness that have max_length constraint
    char_fields = {}
    for field in KYCBusiness._meta.get_fields():
        if isinstance(field, models.CharField) or isinstance(field, models.TextField):
            max_length = getattr(field, 'max_length', None)
            if max_length:
                char_fields[field.name] = max_length
    
    # If a test value is provided, check it against the constraints
    test_results = {}
    if request.method == 'POST':
        for field_name, max_length in char_fields.items():
            field_value = request.POST.get(field_name, '')
            if field_value:
                test_results[field_name] = {
                    'value': field_value,
                    'length': len(field_value),
                    'max_length': max_length,
                    'exceeds': len(field_value) > max_length
                }
    
    return render(request, 'debug_field_lengths.html', {
        'char_fields': char_fields,
        'test_results': test_results
    })

@login_required
def combined_reports(request):
    """
    Combined view for both viewing reports and generating new reports.
    Merges functionality from kyc_reports_list and reports_dashboard.
    """
    # Get filter parameters for reports list
    report_type = request.GET.get('type', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    decision = request.GET.get('decision', '')
    search_query = request.GET.get('q', '')
    
    # Start with all reports
    reports = KYCReport.objects.all()
    
    # Apply filters
    if report_type:
        reports = reports.filter(report_type=report_type)
    
    if date_from:
        try:
            from_date = datetime.strptime(date_from, '%Y-%m-%d')
            reports = reports.filter(generated_at__gte=from_date)
        except ValueError:
            pass
    
    if date_to:
        try:
            to_date = datetime.strptime(date_to, '%Y-%m-%d')
            to_date = to_date.replace(hour=23, minute=59, second=59)
            reports = reports.filter(generated_at__lte=to_date)
        except ValueError:
            pass
    
    if decision:
        reports = reports.filter(decision=decision)
    
    if search_query:
        reports = reports.filter(
            Q(report_id__icontains=search_query) |
            Q(kyc_profile__full_name__icontains=search_query) |
            Q(business_kyc__business_name__icontains=search_query)
        )
    
    # Stats for the dashboard
    total_reports = reports.count()
    approved_reports = reports.filter(decision='APPROVED').count()
    rejected_reports = reports.filter(decision='REJECTED').count()
    individual_reports = reports.filter(report_type='INDIVIDUAL').count()
    business_reports = reports.filter(report_type='BUSINESS').count()
    
    # Get individual profiles with approved or rejected status
    approved_individuals = KYCProfile.objects.filter(
        workflow_state__current_state='APPROVED',
    ).select_related('workflow_state')
    
    rejected_individuals = KYCProfile.objects.filter(
        workflow_state__current_state='REJECTED',
    ).select_related('workflow_state')
    
    # Get business profiles with approved or rejected status
    approved_businesses = KYCBusiness.objects.filter(
        workflow_state__current_state='APPROVED',
    ).select_related('workflow_state')
    
    rejected_businesses = KYCBusiness.objects.filter(
        workflow_state__current_state='REJECTED',
    ).select_related('workflow_state')
    
    # Add risk level information to profiles
    individuals = list(approved_individuals) + list(rejected_individuals)
    businesses = list(approved_businesses) + list(rejected_businesses)
    
    # Get risk levels for individual profiles
    for profile in individuals:
        try:
            test_result = KYCTestResult.objects.filter(kyc_profile=profile).latest('created_at')
            profile.risk_level = test_result.risk_level
        except KYCTestResult.DoesNotExist:
            profile.risk_level = None
    
    # Get risk levels for business profiles
    # Note: It appears KYCTestResult doesn't have a business_kyc field
    # So we'll set risk_level to None for all businesses
    for business in businesses:
        business.risk_level = None
    
    # Get recently generated reports
    recent_reports = KYCReport.objects.all().order_by('-generated_at')[:10]
    
    context = {
        # Reports list context
        'reports': reports,
        'total_reports': total_reports,
        'approved_reports': approved_reports,
        'rejected_reports': rejected_reports,
        'individual_reports': individual_reports,
        'business_reports': business_reports,
        'filter_type': report_type,
        'filter_date_from': date_from,
        'filter_date_to': date_to,
        'filter_decision': decision,
        'search_query': search_query,
        
        # Reports generation context
        'individuals': individuals,
        'businesses': businesses,
        'approved_individuals': approved_individuals.count(),
        'rejected_individuals': rejected_individuals.count(),
        'approved_businesses': approved_businesses.count(),
        'rejected_businesses': rejected_businesses.count(),
        'recent_reports': recent_reports,
    }
    
    return render(request, 'kyc_app/combined_reports.html', context)
